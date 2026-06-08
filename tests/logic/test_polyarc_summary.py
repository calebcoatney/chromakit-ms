"""Unit tests for logic/polyarc_summary.py."""
from __future__ import annotations

from dataclasses import FrozenInstanceError
from pathlib import Path

import openpyxl
import pytest

from logic.polyarc_calibration import Calibration
from logic.polyarc_library import PolyarcLibrary
from logic.polyarc_quantitation import SampleInputs
from logic.polyarc_summary import (
    BatchSummary,
    UnmatchedPeak,
    summarize_batch,
    write_summary_xlsx,
)


FIXTURE_DIR = Path(__file__).parent.parent / 'data'


def _load_minimal_library() -> PolyarcLibrary:
    return PolyarcLibrary.from_csv(FIXTURE_DIR / 'compounds_minimal.csv')


def _load_minimal_calibration() -> Calibration:
    return Calibration.from_csv(FIXTURE_DIR / 'calibration_minimal.csv')


def test_unmatched_peak_is_frozen():
    peak = UnmatchedPeak(
        sample_id='S1', peak_number=3, retention_time=12.5,
        compound_id='Foo', casno='000123-45-6', area=1000.0,
        reason='no_cas_match',
    )
    with pytest.raises(FrozenInstanceError):
        peak.sample_id = 'S2'  # type: ignore[misc]


def test_batch_summary_is_frozen():
    summary = BatchSummary(
        per_sample_peaks={}, per_sample_group_totals={},
        per_sample_inputs={}, unmatched=[], match_stats={},
    )
    with pytest.raises(FrozenInstanceError):
        summary.unmatched = []  # type: ignore[misc]


def test_unmatched_peak_valid_reasons():
    # Document the allowed reasons; not strictly validated at runtime
    # but tests should fail loudly if we typo one anywhere.
    for reason in ('no_cas_match', 'sentinel_cas', 'malformed_cas', 'no_record'):
        UnmatchedPeak(
            sample_id='S1', peak_number=1, retention_time=0.0,
            compound_id='', casno='', area=0.0, reason=reason,
        )


def test_summarize_batch_classifies_unmatched_reasons():
    library = _load_minimal_library()
    calibration = _load_minimal_calibration()
    json_paths = {'S1': FIXTURE_DIR / 'summary_minimal_sample1.json'}
    weights = {'S1': SampleInputs(sample_mass_g=0.1, solvent_mass_g=0.9)}

    summary = summarize_batch(json_paths, weights, library, calibration)

    reasons = {(u.peak_number, u.reason) for u in summary.unmatched if u.sample_id == 'S1'}
    assert (4, 'no_cas_match') in reasons
    assert (5, 'sentinel_cas') in reasons
    assert (6, 'malformed_cas') in reasons
    # Peaks 1, 2, 3 are matched — should NOT appear in unmatched.
    assert not any(u.peak_number in (1, 2, 3) for u in summary.unmatched if u.sample_id == 'S1')


def test_summarize_batch_returns_peakresults_for_matched():
    library = _load_minimal_library()
    calibration = _load_minimal_calibration()
    json_paths = {'S1': FIXTURE_DIR / 'summary_minimal_sample1.json'}
    weights = {'S1': SampleInputs(sample_mass_g=0.1, solvent_mass_g=0.9)}

    summary = summarize_batch(json_paths, weights, library, calibration)

    peaks = summary.per_sample_peaks['S1']
    assert len(peaks) == 6  # all peaks, matched or not
    matched = [p for p in peaks if p.matched]
    assert len(matched) == 3
    assert {p.casno for p in matched} == {'000064-19-7', '000108-95-2', '000498-07-7'}


def test_summarize_batch_group_rollup_sums_wt_pct():
    library = _load_minimal_library()
    calibration = _load_minimal_calibration()
    json_paths = {'S2': FIXTURE_DIR / 'summary_minimal_sample2.json'}
    weights = {'S2': SampleInputs(sample_mass_g=0.1, solvent_mass_g=0.9)}

    summary = summarize_batch(json_paths, weights, library, calibration)

    totals = summary.per_sample_group_totals['S2']
    # Sample 2 has Nonane (Alkane/n-Alkane), Toluene (Aromatic/BTX),
    # Phenol, 2-methyl- (Oxygenate/Phenols/Methylphenol)
    expected_groups = {'Alkane', 'n-Alkane', 'Aromatic', 'BTX',
                       'Oxygenate', 'Phenols', 'Methylphenol',
                       'Total Mass % Accounted'}
    assert expected_groups <= set(totals.keys()), \
        f'Missing groups: {expected_groups - set(totals.keys())}'

    # Pin specific per-peak wt_pct values that a maintainer can hand-verify
    # from the fixture data + library + calibration:
    #
    # For Nonane (sample 2 peak 1): area=15000000, C=9, MW=128.255 (from
    # compounds_minimal.csv), anchor=nonane (known_wt_pct=0.0440433,
    # area=28848616.5, C=9, MW=128.259). DF = (0.1+0.9)/0.1 = 10.
    #   RF = (0.0440433/28848616.5) * (9/9) * (128.255/128.259)
    #   wt_pct = 15000000 * RF * 10  ≈ 0.22900
    assert totals['Alkane'] == pytest.approx(0.22899721751372437, rel=1e-9)
    assert totals['n-Alkane'] == pytest.approx(0.22899721751372437, rel=1e-9)

    # Toluene (peak 2): area=8000000, C=7, MW=92.139, anchor=nonane.
    #   wt_pct = 8000000 * (0.0440433/28848616.5) * (9/7) * (92.139/128.259) * 10
    #          ≈ 0.11281
    assert totals['Aromatic'] == pytest.approx(0.11280821320232672, rel=1e-9)
    assert totals['BTX'] == pytest.approx(0.11280821320232672, rel=1e-9)

    # Phenol, 2-methyl- (peak 3): area=4000000, C=7, MW=108.138, anchor=nonane.
    #   wt_pct = 4000000 * (0.0440433/28848616.5) * (9/7) * (108.138/128.259) * 10
    #          ≈ 0.06620
    assert totals['Methylphenol'] == pytest.approx(0.06619844643620068, rel=1e-9)
    assert totals['Phenols'] == pytest.approx(0.06619844643620068, rel=1e-9)
    assert totals['Oxygenate'] == pytest.approx(0.06619844643620068, rel=1e-9)

    # Total = sum of three peaks' wt_pct ≈ 0.40800
    assert totals['Total Mass % Accounted'] == pytest.approx(0.4080038771522518, rel=1e-9)


def test_summarize_batch_match_stats():
    library = _load_minimal_library()
    calibration = _load_minimal_calibration()
    json_paths = {'S1': FIXTURE_DIR / 'summary_minimal_sample1.json'}
    weights = {'S1': SampleInputs(sample_mass_g=0.1, solvent_mass_g=0.9)}

    summary = summarize_batch(json_paths, weights, library, calibration)

    stats = summary.match_stats['S1']
    assert stats['count_total'] == 6.0
    assert stats['count_matched'] == 3.0
    # Matched peaks: 20M + 10M + 5M = 35M out of 35M + 1M + 0.8M + 0.5M = 37.3M
    expected_area_pct = (20000000 + 10000000 + 5000000) / (20000000 + 10000000 + 5000000 + 1000000 + 800000 + 500000) * 100
    assert stats['area_matched_pct'] == pytest.approx(expected_area_pct, rel=1e-6)


def test_summarize_batch_raises_on_id_mismatch():
    library = _load_minimal_library()
    calibration = _load_minimal_calibration()
    json_paths = {'S1': FIXTURE_DIR / 'summary_minimal_sample1.json'}
    weights = {'S2': SampleInputs(sample_mass_g=0.1, solvent_mass_g=0.9)}

    with pytest.raises(KeyError, match=r"S1.*S2|S2.*S1"):
        summarize_batch(json_paths, weights, library, calibration)


def test_write_summary_xlsx_creates_expected_sheets(tmp_path):
    library = _load_minimal_library()
    calibration = _load_minimal_calibration()
    json_paths = {
        'S1': FIXTURE_DIR / 'summary_minimal_sample1.json',
        'S2': FIXTURE_DIR / 'summary_minimal_sample2.json',
    }
    weights = {
        'S1': SampleInputs(sample_mass_g=0.1, solvent_mass_g=0.9),
        'S2': SampleInputs(sample_mass_g=0.05, solvent_mass_g=0.95),
    }
    summary = summarize_batch(json_paths, weights, library, calibration)

    out_path = tmp_path / 'test_summary.xlsx'
    write_summary_xlsx(summary, out_path)

    assert out_path.exists()
    wb = openpyxl.load_workbook(out_path)
    assert 'Weights' in wb.sheetnames
    assert 'Summary' in wb.sheetnames
    assert 'Sample 1' in wb.sheetnames
    assert 'Sample 2' in wb.sheetnames
    assert 'Unmatched' in wb.sheetnames


def test_write_summary_xlsx_weights_sheet(tmp_path):
    library = _load_minimal_library()
    calibration = _load_minimal_calibration()
    json_paths = {'S1': FIXTURE_DIR / 'summary_minimal_sample1.json'}
    weights = {'S1': SampleInputs(sample_mass_g=0.1, solvent_mass_g=0.9)}
    summary = summarize_batch(json_paths, weights, library, calibration)

    out_path = tmp_path / 'test_summary.xlsx'
    write_summary_xlsx(summary, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb['Weights']
    assert ws['A1'].value == 'Sample ID'
    assert ws['B1'].value == 'Weight'
    assert ws['C1'].value == 'Acetone wt'
    assert ws['D1'].value == 'DF'
    assert ws['A2'].value == 'S1'
    assert ws['B2'].value == pytest.approx(0.1)
    assert ws['C2'].value == pytest.approx(0.9)
    assert ws['D2'].value == pytest.approx(10.0)  # (0.1 + 0.9) / 0.1


def test_write_summary_xlsx_summary_sheet_row_order(tmp_path):
    library = _load_minimal_library()
    calibration = _load_minimal_calibration()
    json_paths = {'S1': FIXTURE_DIR / 'summary_minimal_sample1.json'}
    weights = {'S1': SampleInputs(sample_mass_g=0.1, solvent_mass_g=0.9)}
    summary = summarize_batch(json_paths, weights, library, calibration)

    custom_order = ['Oxygenate', 'Phenols', 'Total Mass % Accounted']
    out_path = tmp_path / 'test_summary.xlsx'
    write_summary_xlsx(summary, out_path, group_row_order=custom_order)

    wb = openpyxl.load_workbook(out_path)
    ws = wb['Summary']
    assert ws['A1'].value is None
    assert ws['B1'].value == 'S1'
    assert ws['A2'].value == 'Oxygenate'
    assert ws['A3'].value == 'Phenols'
    assert ws['A4'].value == 'Total Mass % Accounted'


def test_write_summary_xlsx_unmatched_sheet(tmp_path):
    library = _load_minimal_library()
    calibration = _load_minimal_calibration()
    json_paths = {'S1': FIXTURE_DIR / 'summary_minimal_sample1.json'}
    weights = {'S1': SampleInputs(sample_mass_g=0.1, solvent_mass_g=0.9)}
    summary = summarize_batch(json_paths, weights, library, calibration)

    out_path = tmp_path / 'test_summary.xlsx'
    write_summary_xlsx(summary, out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb['Unmatched']
    assert ws['A1'].value == 'sample_id'
    assert ws['G1'].value == 'reason'
    # 3 unmatched peaks from S1 fixture: rows 2, 3, 4
    reasons = {ws.cell(row=r, column=7).value for r in range(2, 5)}
    assert reasons == {'no_cas_match', 'sentinel_cas', 'malformed_cas'}
