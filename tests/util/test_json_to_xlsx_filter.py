"""Tests for peak-filter helpers in util/json_to_xlsx."""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..'))

from util.json_to_xlsx import _is_unidentified, _is_unquantitated


class TestIsUnidentified:
    def test_none_compound_id_is_unidentified(self):
        assert _is_unidentified({"compound_id": None}) is True

    def test_missing_compound_id_is_unidentified(self):
        assert _is_unidentified({"peak_number": 1}) is True

    def test_empty_string_is_unidentified(self):
        assert _is_unidentified({"compound_id": ""}) is True

    def test_whitespace_only_is_unidentified(self):
        assert _is_unidentified({"compound_id": "   "}) is True

    def test_literal_unknown_is_unidentified(self):
        assert _is_unidentified({"compound_id": "Unknown"}) is True
        assert _is_unidentified({"compound_id": "unknown"}) is True
        assert _is_unidentified({"compound_id": "UNKNOWN"}) is True

    def test_auto_unknown_with_rt_is_unidentified(self):
        """Pattern 'Unknown (X.XXX)' written by integration as placeholder."""
        assert _is_unidentified({"compound_id": "Unknown (5.234)"}) is True
        assert _is_unidentified({"compound_id": "Unknown (12.001)"}) is True

    def test_real_compound_name_is_not_unidentified(self):
        assert _is_unidentified({"compound_id": "Propanoic acid"}) is False
        assert _is_unidentified({"compound_id": "Nonane"}) is False

    def test_compound_id_capitalized_key_also_read(self):
        """JSON sometimes uses 'Compound ID' (notebook-style)."""
        assert _is_unidentified({"Compound ID": "Decane"}) is False
        assert _is_unidentified({"Compound ID": None}) is True


class TestIsUnquantitated:
    def test_no_quantitation_run_returns_false(self):
        """If no peak in the file has wt_percent, filter is a no-op."""
        peaks = [
            {"compound_id": "Decane", "wt_percent": None},
            {"compound_id": "Benzene", "wt_percent": None},
        ]
        assert _is_unquantitated(peaks[0], peaks) is False
        assert _is_unquantitated(peaks[1], peaks) is False

    def test_quantitation_run_unquantitated_peak_returns_true(self):
        """When other peaks have wt_percent, a None means 'was skipped'."""
        peaks = [
            {"compound_id": "Decane", "wt_percent": 50.0},
            {"compound_id": "Mysterium", "wt_percent": None},
        ]
        assert _is_unquantitated(peaks[1], peaks) is True

    def test_quantitation_run_quantitated_peak_returns_false(self):
        peaks = [
            {"compound_id": "Decane", "wt_percent": 50.0},
            {"compound_id": "Benzene", "wt_percent": 30.0},
        ]
        assert _is_unquantitated(peaks[0], peaks) is False
        assert _is_unquantitated(peaks[1], peaks) is False


import json
import os
import tempfile

import pytest
pytest.importorskip('pytestqt')  # ProcessingThread is a QThread

from openpyxl import load_workbook

from util.json_to_xlsx import ProcessingThread, DEFAULT_FORMAT_CONFIG
import copy


def _make_json_file(path: str, peaks: list, sample_id: str = "TEST-001"):
    """Write a minimal valid ChromaKit JSON file for the converter."""
    data = {
        "sample_id": sample_id,
        "timestamp": "2026-06-22 14:00:00",
        "method": "test.chromethod",
        "detector": "FID1",
        "peaks": peaks,
    }
    with open(path, 'w') as f:
        json.dump(data, f)


class TestProcessingThreadFilter:
    def test_skip_unidentified_filters_unknowns_from_xlsx(self, tmp_path):
        # Build a JSON with mixed identified and unidentified peaks
        json_dir = tmp_path / "data"
        json_dir.mkdir()
        json_file = json_dir / "sample.json"
        _make_json_file(str(json_file), peaks=[
            {"peak_number": 1, "retention_time": 2.0, "compound_id": "Decane",
             "area": 1000.0, "width": 0.1, "integrator": "BB",
             "start_time": 1.9, "end_time": 2.1},
            {"peak_number": 2, "retention_time": 3.0, "compound_id": "Unknown",
             "area": 500.0, "width": 0.1, "integrator": "BB",
             "start_time": 2.9, "end_time": 3.1},
            {"peak_number": 3, "retention_time": 4.0, "compound_id": None,
             "area": 200.0, "width": 0.1, "integrator": "BB",
             "start_time": 3.9, "end_time": 4.1},
        ])
        output = tmp_path / "out.xlsx"

        thread = ProcessingThread(
            directory=str(json_dir),
            output_file=str(output),
            format_config=copy.deepcopy(DEFAULT_FORMAT_CONFIG),
            skip_unidentified=True,
        )
        # Call the worker directly (not via run()) to avoid threading
        ok = thread.process_json_to_excel(str(json_dir), str(output))
        assert ok is True
        assert thread._skipped_count == 2

        # Verify the xlsx only has Decane in the peak rows
        wb = load_workbook(str(output))
        ws = wb.active
        # Find rows containing compound names — the exact row depends on header format
        all_values = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
        assert "Decane" in all_values
        assert "Unknown" not in all_values

    def test_skip_unidentified_false_keeps_unknowns(self, tmp_path):
        json_dir = tmp_path / "data"
        json_dir.mkdir()
        json_file = json_dir / "sample.json"
        _make_json_file(str(json_file), peaks=[
            {"peak_number": 1, "retention_time": 2.0, "compound_id": "Decane",
             "area": 1000.0, "width": 0.1, "integrator": "BB",
             "start_time": 1.9, "end_time": 2.1},
            {"peak_number": 2, "retention_time": 3.0, "compound_id": "Unknown",
             "area": 500.0, "width": 0.1, "integrator": "BB",
             "start_time": 2.9, "end_time": 3.1},
        ])
        output = tmp_path / "out.xlsx"

        thread = ProcessingThread(
            directory=str(json_dir),
            output_file=str(output),
            format_config=copy.deepcopy(DEFAULT_FORMAT_CONFIG),
            skip_unidentified=False,
        )
        thread.process_json_to_excel(str(json_dir), str(output))
        assert thread._skipped_count == 0


class TestIsInsideCDataSubtree:
    """Tests for _is_inside_c_data_subtree — skips stale .D-located JSONs.

    Before the .C-folder migration, ChromaKit wrote per-sample JSON results
    inside the .D directory. After the migration, JSONs are written to
    `<sample>.C/results/`. Stale .D-located JSONs may still exist inside
    `<sample>.C/data/<sample>.D/` and would otherwise produce duplicate rows
    in the xlsx output with subtly different sample_ids.
    """
    def test_root_at_results_dir_is_not_inside_c_data(self):
        from util.json_to_xlsx import _is_inside_c_data_subtree
        assert _is_inside_c_data_subtree(
            "/data/2026-06-16/8260-114-01.C/results"
        ) is False

    def test_root_inside_c_data_subtree_is_detected(self):
        from util.json_to_xlsx import _is_inside_c_data_subtree
        assert _is_inside_c_data_subtree(
            "/data/2026-06-16/8260-114-01.C/data/8260-114-01.D"
        ) is True

    def test_root_at_c_folder_itself_is_not_inside_c_data(self):
        from util.json_to_xlsx import _is_inside_c_data_subtree
        assert _is_inside_c_data_subtree("/data/2026-06-16/8260-114-01.C") is False

    def test_plain_directory_is_not_inside_c_data(self):
        from util.json_to_xlsx import _is_inside_c_data_subtree
        assert _is_inside_c_data_subtree("/some/regular/path") is False

    def test_d_folder_outside_c_is_not_inside_c_data(self):
        """Bare .D directories (legacy, not inside a .C container) are NOT skipped."""
        from util.json_to_xlsx import _is_inside_c_data_subtree
        assert _is_inside_c_data_subtree("/data/2026-06-16/8260-114-01.D") is False


class TestProcessingThreadSkipsStaleCData:
    """Integration test: when a .C folder has both a new results/ JSON AND a
    stale data/<sample>.D/ JSON, the converter writes ONLY the new one."""

    def test_stale_d_json_inside_c_folder_is_skipped(self, tmp_path):
        # Build the directory tree:
        #   sample.C/
        #     results/sample - FID1A.json   (new, sample_id='sample')
        #     data/sample.D/sample - FID1A.json  (stale, sample_id='sample.D')
        c_folder = tmp_path / "sample.C"
        results_dir = c_folder / "results"
        data_d_dir = c_folder / "data" / "sample.D"
        results_dir.mkdir(parents=True)
        data_d_dir.mkdir(parents=True)

        _make_json_file(
            str(results_dir / "sample - FID1A.json"),
            peaks=[{"peak_number": 1, "retention_time": 2.0, "compound_id": "Decane",
                    "area": 1000.0, "width": 0.1, "integrator": "BB",
                    "start_time": 1.9, "end_time": 2.1}],
            sample_id="sample",
        )
        _make_json_file(
            str(data_d_dir / "sample - FID1A.json"),
            peaks=[{"peak_number": 1, "retention_time": 2.0, "compound_id": "Hexane",
                    "area": 999.0, "width": 0.1, "integrator": "BB",
                    "start_time": 1.9, "end_time": 2.1}],
            sample_id="sample.D",  # legacy sample_id
        )

        output = tmp_path / "out.xlsx"
        thread = ProcessingThread(
            directory=str(tmp_path),
            output_file=str(output),
            format_config=copy.deepcopy(DEFAULT_FORMAT_CONFIG),
            skip_unidentified=False,
        )
        ok = thread.process_json_to_excel(str(tmp_path), str(output))
        assert ok is True

        wb = load_workbook(str(output))
        ws = wb.active
        all_values = [str(c.value) for row in ws.iter_rows() for c in row
                      if c.value is not None]
        # Only the new file should be present
        assert "Decane" in all_values
        assert "Hexane" not in all_values
        assert "sample" in all_values
        assert "sample.D" not in all_values
