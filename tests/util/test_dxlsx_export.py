import numpy as np
import pytest
from openpyxl import load_workbook
from util import dxlsx_export as dx


class StubDataFile:
    def __init__(self, name, xlabels, data, metadata=None):
        self.name = name
        self.xlabels = np.asarray(xlabels)
        self.data = np.asarray(data)
        self.metadata = metadata or {}


class StubDataDir:
    def __init__(self, name, datafiles, metadata=None):
        self.name = name
        self.datafiles = datafiles
        self.metadata = metadata or {}

    def get_file(self, fname):
        for f in self.datafiles:
            if f.name == fname:
                return f
        raise KeyError(fname)


def make_dir():
    fid = StubDataFile("FID1A.ch", [0.0, 0.5, 1.0], [10.0, 20.0, 30.0],
                       metadata={"notebook": "SampleA"})
    ms = StubDataFile("data.ms", [0.5, 1.0], [[1, 2, 3], [10, 20, 30]])
    return StubDataDir("dirname", [fid, ms], metadata={"notebook": "DirNB"})


def test_list_signals_returns_ch_and_ms():
    d = make_dir()
    assert dx.list_signals(d) == ["FID1A.ch", "data.ms"]


def test_read_signal_ch():
    d = make_dir()
    x, y = dx.read_signal(d, "FID1A.ch")
    assert np.allclose(x, [0.0, 0.5, 1.0])
    assert np.allclose(y, [10.0, 20.0, 30.0])


def test_read_signal_ms_is_tic():
    d = make_dir()
    x, y = dx.read_signal(d, "data.ms")
    assert np.allclose(x, [0.5, 1.0])
    assert np.allclose(y, [6.0, 60.0])


def test_read_notebook_prefers_detector_file():
    d = make_dir()
    assert dx.read_notebook(d, "/some/path.D") == "SampleA"


def test_read_notebook_falls_back_to_dir_then_basename():
    fid = StubDataFile("FID1A.ch", [0.0], [1.0], metadata={})
    d = StubDataDir("dname", [fid], metadata={})
    assert dx.read_notebook(d, "/x/basename.D") == "dname"


def test_read_notebook_uses_basename_when_no_name():
    fid = StubDataFile("FID1A.ch", [0.0], [1.0], metadata={})
    d = StubDataDir(None, [fid], metadata={})  # name=None
    assert dx.read_notebook(d, "/data/MySample.D") == "MySample"


def test_read_notebook_basename_handles_trailing_slash():
    fid = StubDataFile("FID1A.ch", [0.0], [1.0], metadata={})
    d = StubDataDir(None, [fid], metadata={})  # name=None
    assert dx.read_notebook(d, "/data/MySample.D/") == "MySample"


def test_build_grid_union_range():
    sig_x = {"A": np.array([0.0, 1.0]), "B": np.array([0.5, 2.0])}
    grid = dx.build_time_grid(sig_x, skip_solvent_delay=False, has_ms=False,
                              ms_x=None, n=5)
    assert np.isclose(grid[0], 0.0)
    assert np.isclose(grid[-1], 2.0)
    assert len(grid) == 5


def test_build_grid_clips_to_ms_start_when_enabled():
    sig_x = {"FID1A.ch": np.array([0.0, 2.0]), "data.ms": np.array([1.8, 2.0])}
    grid = dx.build_time_grid(sig_x, skip_solvent_delay=True, has_ms=True,
                              ms_x=np.array([1.8, 2.0]), n=5)
    assert np.isclose(grid[0], 1.8)
    assert np.isclose(grid[-1], 2.0)


def test_build_grid_raises_on_empty():
    with pytest.raises(ValueError):
        dx.build_time_grid({}, skip_solvent_delay=False, has_ms=False, ms_x=None, n=5)


def test_build_grid_no_clip_when_flag_off_even_with_ms():
    sig_x = {"FID1A.ch": np.array([0.0, 2.0]), "data.ms": np.array([1.8, 2.0])}
    grid = dx.build_time_grid(sig_x, skip_solvent_delay=False, has_ms=True,
                              ms_x=np.array([1.8, 2.0]), n=5)
    assert np.isclose(grid[0], 0.0)


def test_resample_masks_out_of_range_to_nan():
    grid = np.array([0.0, 0.5, 1.0, 1.5, 2.0])
    x = np.array([0.5, 1.0, 1.5])
    y = np.array([5.0, 10.0, 15.0])
    out = dx.resample_to_grid(grid, x, y)
    assert np.isnan(out[0])
    assert np.isclose(out[1], 5.0)
    assert np.isclose(out[2], 10.0)
    assert np.isclose(out[3], 15.0)
    assert np.isnan(out[4])


def test_sanitize_removes_invalid_and_truncates():
    used = set()
    name = dx.safe_sheet_name("a/b:c*d?e[f]g" * 5, used)
    for ch in r'[]:*?/\\':
        assert ch not in name
    assert len(name) <= 31
    assert name in used


def test_sanitize_dedupes_collisions():
    used = set()
    n1 = dx.safe_sheet_name("Sample", used)
    n2 = dx.safe_sheet_name("Sample", used)
    n3 = dx.safe_sheet_name("Sample", used)
    assert n1 == "Sample"
    assert n2 == "Sample_2"
    assert n3 == "Sample_3"


def test_sanitize_empty_falls_back():
    used = set()
    name = dx.safe_sheet_name("", used)
    assert name == "Sheet"


def test_sanitize_no_trailing_space_after_truncation():
    used = set()
    name = dx.safe_sheet_name("X" * 30 + "   tail", used)
    assert len(name) <= 31
    assert name == name.strip()  # no leading/trailing whitespace


def test_sanitize_long_base_with_many_collisions_stay_unique_and_bounded():
    used = set()
    base = "N" * 31
    names = [dx.safe_sheet_name(base, used) for _ in range(15)]
    assert len(set(names)) == 15          # all unique
    assert all(len(n) <= 31 for n in names)
    assert all(n == n.strip() for n in names)


def test_build_sheet_rows_shape_and_headers():
    d = make_dir()  # FID1A.ch (0..1), data.ms (0.5..1.0)
    header, rows = dx.build_sheet_rows(
        d, selected=["FID1A.ch", "data.ms"],
        skip_solvent_delay=False, n=4)
    assert header == ["Time (min)", "FID1A.ch", "data.ms"]
    assert len(rows) == 4
    assert len(rows[0]) == 3
    # first grid point 0.0 is before MS start -> MS cell is None
    assert rows[0][2] is None


def test_build_sheet_rows_skips_absent_signals():
    fid = StubDataFile("FID1A.ch", [0.0, 1.0], [1.0, 2.0])
    d = StubDataDir("d", [fid])
    header, rows = dx.build_sheet_rows(
        d, selected=["FID1A.ch", "data.ms"],
        skip_solvent_delay=True, n=3)
    # data.ms not present -> column omitted
    assert header == ["Time (min)", "FID1A.ch"]


def test_build_sheet_rows_empty_when_no_signals_present():
    fid = StubDataFile("FID1A.ch", [0.0, 1.0], [1.0, 2.0])
    d = StubDataDir("d", [fid])
    header, rows = dx.build_sheet_rows(
        d, selected=["data.ms"],  # not present in this folder
        skip_solvent_delay=False, n=5)
    assert header == ["Time (min)"]
    assert rows == []


def test_write_workbook_creates_sheets(tmp_path):
    out = tmp_path / "out.xlsx"
    sheets = [
        ("SampleA", ["Time (min)", "FID1A.ch"], [[0.0, 1.0], [0.5, None]]),
        ("SampleB", ["Time (min)", "data.ms"], [[1.0, 6.0]]),
    ]
    dx.write_workbook(str(out), sheets)
    wb = load_workbook(str(out))
    assert wb.sheetnames == ["SampleA", "SampleB"]
    ws = wb["SampleA"]
    assert ws.cell(row=1, column=1).value == "Time (min)"
    assert ws.cell(row=1, column=1).font.bold is True
    assert ws.cell(row=3, column=2).value is None  # NaN cell empty


def test_write_workbook_raises_on_no_sheets(tmp_path):
    out = tmp_path / "empty.xlsx"
    with pytest.raises(ValueError):
        dx.write_workbook(str(out), [])


def test_export_folders_builds_sheets_and_skips_failures(tmp_path):
    good = make_dir()

    def reader(path):
        if path == "GOOD":
            return good
        raise IOError("bad .D")

    out = tmp_path / "wb.xlsx"
    logs = []
    result = dx.export_folders(
        ["GOOD", "BAD"], selected=["FID1A.ch", "data.ms"],
        skip_solvent_delay=False, n=4, out_path=str(out),
        reader=reader, log=logs.append, progress=lambda i: None)

    assert result["exported"] == 1
    assert result["skipped"] == 1
    assert out.exists()
    assert any("bad .D" in m or "BAD" in m for m in logs)


def test_export_folders_all_fail_writes_no_file(tmp_path):
    def reader(path):
        raise IOError("boom")

    out = tmp_path / "none.xlsx"
    logs = []
    result = dx.export_folders(
        ["A", "B"], selected=["FID1A.ch"], skip_solvent_delay=False, n=4,
        out_path=str(out), reader=reader, log=logs.append)
    assert result == {"exported": 0, "skipped": 2}
    assert not out.exists()
    assert any("No sheets to write." in m for m in logs)


def test_export_folders_empty_folder_list(tmp_path):
    out = tmp_path / "none.xlsx"
    result = dx.export_folders(
        [], selected=["FID1A.ch"], skip_solvent_delay=False, n=4,
        out_path=str(out), reader=lambda p: None, log=lambda m: None)
    assert result == {"exported": 0, "skipped": 0}
    assert not out.exists()


def test_export_folders_skips_folder_with_no_selected_signals(tmp_path):
    fid_only = StubDataDir("d", [StubDataFile("FID1A.ch", [0.0, 1.0], [1.0, 2.0])])

    def reader(path):
        return fid_only

    out = tmp_path / "wb.xlsx"
    logs = []
    result = dx.export_folders(
        ["X"], selected=["data.ms"],  # not present -> len(header)<=1 skip
        skip_solvent_delay=False, n=4, out_path=str(out),
        reader=reader, log=logs.append)
    assert result == {"exported": 0, "skipped": 1}
    assert not out.exists()
    assert any("no selected signals present" in m for m in logs)


def test_export_folders_progress_reaches_100(tmp_path):
    good = make_dir()
    out = tmp_path / "wb.xlsx"
    seen = []
    dx.export_folders(
        ["G"], selected=["FID1A.ch", "data.ms"], skip_solvent_delay=False,
        n=4, out_path=str(out), reader=lambda p: good,
        log=lambda m: None, progress=seen.append)
    assert seen and seen[-1] == 100
