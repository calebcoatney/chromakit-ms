"""
One-off extraction script: reads Kelly's GCMS-Polyarc workbook and writes
data/polyarc/compounds.csv.

Each row of the workbook's "Compounds" sheet describes one library compound.
We detect the anchor standard by looking at the column-L (Polyarc RF) formula:
that formula divides by the anchor's known_wt_pct cell, which is `$P$2` for
nonane and `$P$6` for levoglucosan in Kelly's layout. We classify the row
by which cell reference appears in the formula.

Usage:
    python data/polyarc/extract_from_xlsx.py \\
        ~/Library/CloudStorage/OneDrive-NREL/GC-MS/examples/GCMS-Polyarc_Alder\\ FPO-11_12_March\\ 2026\\ \\(1\\).xlsx \\
        data/polyarc/compounds.csv

Re-run only when Kelly publishes a new library version.
"""
import argparse
import csv
from pathlib import Path

import openpyxl


def _atom_count(value, atom_label, row_num):
    """Convert an atom-count cell to int; empty/None → ''.

    Raises ValueError on non-integer float (e.g. 8.5) — these would silently
    corrupt downstream stoichiometry.
    """
    if value is None or value == '':
        return ''
    if not isinstance(value, (int, float)):
        return ''
    if isinstance(value, float) and not value.is_integer():
        raise ValueError(
            f'Row {row_num}: atom count {atom_label}={value!r} is a '
            f'non-integer float; refusing to silently truncate.'
        )
    return int(value)


def extract(src_xlsx: Path, dst_csv: Path) -> int:
    """Extract Compounds sheet to CSV. Returns row count written."""
    # Need both views: data_only=False gives us the column-L formula text we
    # use to detect the anchor (which anchor's cell P2/P6 the row points at);
    # data_only=True gives us the cached numeric values for everything else.
    wb_f = openpyxl.load_workbook(src_xlsx, data_only=False)
    wb_d = openpyxl.load_workbook(src_xlsx, data_only=True)

    if 'Compounds' not in wb_f.sheetnames:
        raise ValueError(
            f"Source workbook {src_xlsx} has no 'Compounds' sheet. "
            f"Available sheets: {wb_f.sheetnames}"
        )

    ws_formula = wb_f['Compounds']
    ws_values = wb_d['Compounds']

    rows = []
    for r in range(2, ws_formula.max_row + 1):
        name = ws_values.cell(r, 1).value
        if name is None:
            continue
        cas = ws_values.cell(r, 2).value
        g1 = ws_values.cell(r, 3).value
        g2 = ws_values.cell(r, 4).value
        g3 = ws_values.cell(r, 5).value
        C = ws_values.cell(r, 6).value
        H = ws_values.cell(r, 7).value
        O = ws_values.cell(r, 8).value
        S = ws_values.cell(r, 9).value
        N = ws_values.cell(r, 10).value
        MW = ws_values.cell(r, 11).value

        L_formula = ws_formula.cell(r, 12).value
        if isinstance(L_formula, str) and '$P$2' in L_formula:
            anchor = 'nonane'
        elif isinstance(L_formula, str) and '$P$6' in L_formula:
            anchor = 'levoglucosan'
        else:
            anchor = ''

        rows.append({
            'compound': str(name).strip(),
            'cas': str(cas).strip() if cas is not None else '',
            'group1': g1 or '',
            'group2': g2 or '',
            'group3': g3 or '',
            'C': _atom_count(C, 'C', r),
            'H': _atom_count(H, 'H', r),
            'O': _atom_count(O, 'O', r),
            'S': _atom_count(S, 'S', r),
            'N': _atom_count(N, 'N', r),
            'MW': round(MW, 4) if isinstance(MW, (int, float)) else '',
            'anchor': anchor,
        })

    fieldnames = ['compound', 'cas', 'group1', 'group2', 'group3',
                  'C', 'H', 'O', 'S', 'N', 'MW', 'anchor']
    with open(dst_csv, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)

    return len(rows)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument('src_xlsx', type=Path, help="Kelly's GCMS-Polyarc workbook")
    ap.add_argument('dst_csv', type=Path, help='Output compounds.csv path')
    args = ap.parse_args()

    args.dst_csv.parent.mkdir(parents=True, exist_ok=True)
    n = extract(args.src_xlsx, args.dst_csv)
    print(f'Wrote {n} rows to {args.dst_csv}')

    # Surface any rows whose column-L formula didn't match a known anchor cell.
    # An '' anchor silently breaks downstream quantitation (compound becomes
    # orphaned from any calibration point).
    import csv as _csv
    with open(args.dst_csv, newline='', encoding='utf-8') as f:
        no_anchor = sum(1 for row in _csv.DictReader(f) if not row['anchor'])
    if no_anchor:
        print(f'WARNING: {no_anchor} rows have empty anchor (column L formula did '
              f'not reference $P$2 or $P$6). These compounds will be unquantifiable.')
    else:
        print('All rows have an anchor.')


if __name__ == '__main__':
    main()
