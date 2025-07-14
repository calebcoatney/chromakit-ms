# -*- coding: utf-8 -*-
"""
Created on Thu Dec  5 11:28:40 2024

@author: ccoatney
"""

import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font
import tkinter as tk
from tkinter import filedialog


def process_json_to_excel(directory, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Chromatograms"

    starting_row = 1  # Start row for writing each file

    for root, _, files in os.walk(directory):
        # Filter for .json files in the current folder
        json_files = [f for f in files if f.endswith('.json')]

        if not json_files:
            continue  # Skip folders without .json files

        for file in json_files:
            json_path = os.path.join(root, file)
            try:
                with open(json_path, 'r') as f:
                    json_data = json.load(f)

                # Write header data
                headers = [
                    ("Sample ID:", json_data.get("sample_id", "")),
                    ("Timestamp:", json_data.get("timestamp", "")),
                    ("Method:", json_data.get("method", "")),
                    ("Detector:", json_data.get("detector", "")),
                ]

                current_row = starting_row  # Start at the current row
                col_offset = 1  # Always start at column 1

                for header, value in headers:
                    cell = ws.cell(row=current_row, column=col_offset, value=header)
                    cell.font = Font(bold=True)
                    ws.cell(row=current_row, column=col_offset + 1, value=value)
                    current_row += 1

                # Optional signal and notebook fields
                ws.cell(row=current_row, column=col_offset, value=json_data.get("signal", ""))
                current_row += 1
                ws.cell(row=current_row, column=col_offset, value=json_data.get("notebook", ""))
                current_row += 2

                # Write peaks table header
                peaks_headers = ["Peak #", "Ret Time", "Integrator",
                                 "Width", "Area", "Start Time", "End Time"]
                for col, header in enumerate(peaks_headers, start=col_offset):
                    cell = ws.cell(row=current_row, column=col, value=header)

                current_row += 1

                # Write peaks data
                for peak in json_data.get("peaks", []):
                    ws.cell(row=current_row, column=col_offset,
                            value=int(peak.get("peak_number", 0)))
                    ws.cell(row=current_row, column=col_offset + 1,
                            value=float(peak.get("retention_time", 0.0)))
                    ws.cell(row=current_row, column=col_offset + 2, value=peak.get("integrator", ""))
                    ws.cell(row=current_row, column=col_offset + 3, value=float(peak.get("width", 0.0)))
                    ws.cell(row=current_row, column=col_offset + 4, value=float(peak.get("area", 0.0)))
                    ws.cell(row=current_row, column=col_offset + 5, value=float(peak.get("start_time", 0.0)))
                    ws.cell(row=current_row, column=col_offset + 6, value=float(peak.get("end_time", 0.0)))
                    current_row += 1

                # After writing this file, update starting_row for the next file
                starting_row = current_row + 2  # Add a gap between files

            except Exception as e:
                print(f"Error reading {json_path}: {e}")

    wb.save(output_file)
    print(f"Data successfully written to {output_file}")


# Initialize Tkinter root window (it won't show on the screen)
root = tk.Tk()
root.withdraw()  # Hide the main Tkinter window

# Ask the user to select the folder containing the JSON files
directory_to_search = filedialog.askdirectory(title="Select Folder to Search for .json Files")
if not directory_to_search:
    print("No directory selected. Exiting.")
    exit()

# Ask the user to select the output Excel file location
output_excel_file = filedialog.asksaveasfilename(
    title="Save Excel File", defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
if not output_excel_file:
    print("No output file selected. Exiting.")
    exit()

# Process the selected directory and save the results
process_json_to_excel(directory_to_search, output_excel_file)
