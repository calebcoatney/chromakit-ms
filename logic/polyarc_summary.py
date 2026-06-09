"""
Polyarc batch summary: roll up per-sample peak results into Kelly-style group totals.

Pure Python (stdlib + openpyxl). No imports from `ui/` or `api/`.

See docs/superpowers/specs/2026-06-08-polyarc-batch-summary-and-validation-design.md.
"""
from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook

from logic.polyarc_calibration import Calibration
from logic.polyarc_library import PolyarcLibrary
from logic.polyarc_quantitation import PeakResult, SampleInputs, quantitate


_SENTINEL_CAS_NORMALIZED = '000000-00-0'

_SAMPLE_PEAK_HEADERS = [
    'compound_id', 'casno', 'retention_time', 'FID_area',
    'MW', 'C', 'RF', 'wt_pct', 'mass_mg', 'mol', 'mol_C',
    'group1', 'group2', 'group3',
]


@dataclass(frozen=True)
class UnmatchedPeak:
    """A peak that could not be quantitated.

    reason ∈ {'no_cas_match', 'sentinel_cas', 'malformed_cas', 'no_record'}.
    """
    sample_id: str
    peak_number: int
    retention_time: float
    compound_id: str
    casno: str
    area: float
    reason: str


@dataclass(frozen=True)
class BatchSummary:
    """Aggregated quantitation results for a batch of samples."""
    per_sample_peaks: dict[str, list[PeakResult]]
    per_sample_group_totals: dict[str, dict[str, float]]
    per_sample_inputs: dict[str, SampleInputs]
    unmatched: list[UnmatchedPeak]
    match_stats: dict[str, dict[str, float]]


def _normalize_cas(cas: str) -> tuple[str, bool]:
    """Return (normalized_cas, is_malformed). Mirrors PolyarcLibrary._pad_cas."""
    if not cas:
        return ('', False)
    cas = cas.strip()
    parts = cas.split('-')
    if len(parts) != 3:
        return (cas, True)
    try:
        normalized = f'{int(parts[0]):06d}-{parts[1]}-{parts[2]}'
    except ValueError:
        return (cas, True)
    return (normalized, False)


def _classify_unmatched(peak: dict, library: PolyarcLibrary) -> str:
    """Bucket an unmatched peak by why the library lookup failed.

    Caller must have already confirmed `quantitate()` returned an unmatched
    PeakResult for this peak (i.e., both CAS and compound_id lookups failed,
    OR the matching record had C=0).

    Buckets, in priority order:
      - 'malformed_cas': CAS does not parse as 'NNNNNN-NN-N' (3 hyphenated
        integer segments). The library cannot be queried.
      - 'sentinel_cas': CAS is the sentinel '000000-00-0' that the library
        deliberately excludes from CAS indexing.
      - 'no_record': CAS field is empty/missing AND the compound_id did not
        resolve by name either.
      - 'no_cas_match': CAS is well-formed but absent from the library, and
        the compound_id also did not resolve by name.
    """
    casno = peak.get('casno', '')
    normalized, is_malformed = _normalize_cas(casno)
    if is_malformed:
        return 'malformed_cas'
    if normalized == _SENTINEL_CAS_NORMALIZED:
        return 'sentinel_cas'
    # CAS was empty or missing — peak reached this point because lookup also
    # failed by name. Treated as 'no_record' (no library entry of any kind).
    if not normalized:
        return 'no_record'
    # CAS was well-formed and not the sentinel, but neither it nor the
    # compound_id resolved to a library record.
    return 'no_cas_match'


def summarize_batch(
    json_paths: dict[str, Path],
    weights: dict[str, SampleInputs],
    library: PolyarcLibrary,
    calibration: Calibration,
) -> BatchSummary:
    """Quantitate every peak in every sample, roll up group totals, collect unmatched.

    Args:
        json_paths: sample_id -> path to chromakit *-FID1A.json file
        weights: sample_id -> SampleInputs (sample mass + solvent mass)
        library: indexed compound library
        calibration: anchor calibration points

    Returns:
        BatchSummary with per-sample peaks, per-sample group totals,
        list of unmatched peaks, and per-sample match statistics.

    Raises:
        KeyError: if a sample_id appears in json_paths but not in weights,
            or vice versa.
    """
    json_ids = set(json_paths.keys())
    weight_ids = set(weights.keys())
    if json_ids != weight_ids:
        only_in_json = json_ids - weight_ids
        only_in_weights = weight_ids - json_ids
        raise KeyError(
            f'json_paths and weights have different sample IDs. '
            f'Only in json_paths: {sorted(only_in_json)}. '
            f'Only in weights: {sorted(only_in_weights)}.'
        )

    per_sample_peaks: dict[str, list[PeakResult]] = {}
    per_sample_group_totals: dict[str, dict[str, float]] = {}
    per_sample_inputs: dict[str, SampleInputs] = {}
    unmatched: list[UnmatchedPeak] = []
    match_stats: dict[str, dict[str, float]] = {}

    for sample_id, json_path in json_paths.items():
        sample = weights[sample_id]
        per_sample_inputs[sample_id] = sample

        with open(json_path, encoding='utf-8') as f:
            data = json.load(f)
        raw_peaks = data.get('peaks', [])

        results = quantitate(raw_peaks, sample, library, calibration)
        per_sample_peaks[sample_id] = results

        # Classify unmatched
        for peak_input, result in zip(raw_peaks, results):
            if not result.matched:
                reason = _classify_unmatched(peak_input, library)
                unmatched.append(UnmatchedPeak(
                    sample_id=sample_id,
                    peak_number=peak_input.get('peak_number', 0),
                    retention_time=float(peak_input.get('retention_time', 0.0)),
                    compound_id=peak_input.get('compound_id', ''),
                    casno=peak_input.get('casno', ''),
                    area=float(peak_input.get('area', 0.0)),
                    reason=reason,
                ))

        # Group rollup
        group_totals: dict[str, float] = {}
        for r in results:
            if not r.matched or r.wt_pct is None or r.record is None:
                continue
            for g in (r.record.group1, r.record.group2, r.record.group3):
                if g and g != '0':  # Kelly's library uses '0' as a placeholder
                    group_totals[g] = group_totals.get(g, 0.0) + r.wt_pct
        group_totals['Total Mass % Accounted'] = sum(
            r.wt_pct for r in results
            if r.matched and r.wt_pct is not None
        )
        per_sample_group_totals[sample_id] = group_totals

        # Match stats
        total_peaks = len(results)
        matched_count = sum(1 for r in results if r.matched)
        total_area = sum(r.area for r in results) or 1.0  # avoid divide-by-zero
        matched_area = sum(r.area for r in results if r.matched)
        match_stats[sample_id] = {
            'count_matched': float(matched_count),
            'count_total': float(total_peaks),
            'area_matched_pct': matched_area / total_area * 100.0,
        }

    return BatchSummary(
        per_sample_peaks=per_sample_peaks,
        per_sample_group_totals=per_sample_group_totals,
        per_sample_inputs=per_sample_inputs,
        unmatched=unmatched,
        match_stats=match_stats,
    )


def write_summary_xlsx(
    summary: BatchSummary,
    out_path: Path,
    *,
    group_row_order: list[str] | None = None,
) -> None:
    """Write a Kelly-style summary workbook from a BatchSummary.

    Sheets:
        Weights  — sample ID, weight, acetone, DF
        Summary  — group totals: columns = samples, rows = groups
        Sample N — per-peak detail (one sheet per sample, in iteration order)
        Unmatched — all UnmatchedPeak rows

    Args:
        summary: BatchSummary produced by summarize_batch()
        out_path: file path to write
        group_row_order: optional list of group names to use as the Summary
            sheet row order. If None, uses sorted union of all observed groups
            with 'Total Mass % Accounted' first.
    """
    wb = Workbook()
    # Workbook() creates a default sheet; remove it.
    default = wb.active
    wb.remove(default)

    sample_ids = list(summary.per_sample_inputs.keys())

    # Weights sheet
    ws = wb.create_sheet('Weights')
    ws.append(['Sample ID', 'Weight', 'Acetone wt', 'DF'])
    for sid in sample_ids:
        s = summary.per_sample_inputs[sid]
        ws.append([sid, s.sample_mass_g, s.solvent_mass_g, s.dilution_factor])

    # Summary sheet
    ws = wb.create_sheet('Summary')
    ws.append([None, *sample_ids])
    if group_row_order is None:
        all_groups: set[str] = set()
        for totals in summary.per_sample_group_totals.values():
            all_groups.update(totals.keys())
        # 'Total Mass % Accounted' first, then sorted rest
        rest = sorted(all_groups - {'Total Mass % Accounted'})
        group_row_order = (['Total Mass % Accounted'] if 'Total Mass % Accounted' in all_groups else []) + rest
    for group in group_row_order:
        row = [group]
        for sid in sample_ids:
            row.append(summary.per_sample_group_totals.get(sid, {}).get(group, 0.0))
        ws.append(row)

    # Sample N sheets (one per sample, in iteration order)
    for n, sid in enumerate(sample_ids, start=1):
        ws = wb.create_sheet(f'Sample {n}')
        ws['A1'] = sid
        # Headers explicitly on row 2 so row 1 holds only the sample ID
        for col_idx, header in enumerate(_SAMPLE_PEAK_HEADERS, start=1):
            ws.cell(row=2, column=col_idx, value=header)
        for r in summary.per_sample_peaks[sid]:
            if not r.matched or r.record is None:
                continue
            ws.append([
                r.compound_id,
                str(r.casno),  # text to preserve leading zeros
                r.retention_time,
                r.area,
                r.record.MW,
                r.record.C,
                r.RF,
                r.wt_pct,
                r.mass_mg,
                r.mol,
                r.mol_C,
                r.record.group1,
                r.record.group2,
                r.record.group3,
            ])

    # Unmatched sheet
    ws = wb.create_sheet('Unmatched')
    ws.append(['sample_id', 'peak_number', 'retention_time',
               'compound_id', 'casno', 'area', 'reason'])
    for u in summary.unmatched:
        ws.append([u.sample_id, u.peak_number, u.retention_time,
                   u.compound_id, str(u.casno), u.area, u.reason])

    wb.save(out_path)
