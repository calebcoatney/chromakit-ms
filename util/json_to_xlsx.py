# -*- coding: utf-8 -*-
"""
ChromaKit-MS JSON ↔ Excel Converter

A utility to convert between ChromaKit-MS integration results and Excel format:
- JSON → Excel: Compile integration results from multiple .D directories into a single Excel summary
- Excel → JSON: Update JSON files with compound IDs from manually annotated Excel files

Created on Thu Dec  5 11:28:40 2024
Updated on Jul 28, 2025 - Converted to PySide6 GUI and added compound ID support
Updated on Aug  6, 2025 - Added bidirectional functionality for updating JSON from Excel

@author: ccoatney
"""

import os
import sys
import json
import copy
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from PySide6.QtWidgets import (
    QApplication, QDialog, QVBoxLayout, QHBoxLayout, QWidget,
    QPushButton, QLabel, QFileDialog, QTextEdit, QProgressBar,
    QMessageBox, QGroupBox, QCheckBox, QTabWidget, QLineEdit,
    QScrollArea, QComboBox, QInputDialog, QSizePolicy, QFormLayout,
    QListWidget, QListWidgetItem, QAbstractItemView
)
from PySide6.QtCore import Qt, QThread, Signal, QSettings, QSize
from PySide6.QtGui import QFont


# ---------------------------------------------------------------------------
# Default format configuration
# ---------------------------------------------------------------------------

DEFAULT_FORMAT_CONFIG = {
    'header_fields': [
        {'key': 'sample_id',  'label': 'Sample ID:',  'enabled': True,  'custom_label': '', 'show_label': True},
        {'key': 'timestamp',  'label': 'Timestamp:',  'enabled': True,  'custom_label': '', 'show_label': True},
        {'key': 'method',     'label': 'Method:',     'enabled': True,  'custom_label': '', 'show_label': True},
        {'key': 'detector',   'label': 'Detector:',   'enabled': True,  'custom_label': '', 'show_label': True},
        {'key': 'signal',     'label': 'Signal:',     'enabled': True,  'custom_label': '', 'show_label': True},
        {'key': 'notebook',   'label': 'Notebook:',   'enabled': True,  'custom_label': '', 'show_label': True},
    ],
    'timestamp_format': 'as_stored',
    'columns': [
        {'key': 'Compound ID',    'display': 'Compound ID',   'enabled': True,  'custom_name': ''},
        {'key': 'peak_number',    'display': 'Peak #',        'enabled': True,  'custom_name': ''},
        {'key': 'retention_time', 'display': 'Ret Time',      'enabled': True,  'custom_name': ''},
        {'key': 'integrator',     'display': 'Integrator',    'enabled': True,  'custom_name': ''},
        {'key': 'width',          'display': 'Width',         'enabled': True,  'custom_name': ''},
        {'key': 'area',           'display': 'Area',          'enabled': True,  'custom_name': ''},
        {'key': 'start_time',     'display': 'Start Time',    'enabled': True,  'custom_name': ''},
        {'key': 'end_time',       'display': 'End Time',      'enabled': True,  'custom_name': ''},
        {'key': 'Qual',           'display': 'Match Score',   'enabled': True,  'custom_name': ''},
        {'key': 'casno',          'display': 'CAS No',        'enabled': True,  'custom_name': ''},
        {'key': 'quality_issues', 'display': 'Quality Issues','enabled': True,  'custom_name': ''},
        {'key': 'mol_C',          'display': 'mol C',         'enabled': False, 'custom_name': ''},
        {'key': 'mol_C_percent',  'display': 'mol C %',       'enabled': False, 'custom_name': ''},
        {'key': 'num_carbons',    'display': '# Carbons',     'enabled': False, 'custom_name': ''},
        {'key': 'mol',            'display': 'mol',           'enabled': False, 'custom_name': ''},
        {'key': 'mass_mg',        'display': 'mass (mg)',     'enabled': False, 'custom_name': ''},
        {'key': 'mol_percent',    'display': 'mol %',         'enabled': False, 'custom_name': ''},
        {'key': 'wt_percent',     'display': 'wt %',          'enabled': False, 'custom_name': ''},
    ],
}

# Timestamp parse patterns to try when reformatting
_TS_PARSE_FMTS = [
    "%d %b %y  %I:%M %p",
    "%d %b %y %I:%M %p",
    "%d %b %Y %I:%M %p",
    "%d-%b-%y %H:%M:%S",
    "%m/%d/%Y %H:%M:%S",
    "%Y-%m-%dT%H:%M:%S",
    "%d %b %y  %H:%M",
    "%d %b %y %H:%M",
]

# Human-readable label → Excel number-format string (or special token "as_stored").
# Excel format strings are used directly as cell.number_format, giving native date cells.
# Any entry whose format value starts with "%" is treated as a strftime fallback (text cell).
TIMESTAMP_FORMAT_OPTIONS = [
    ("As stored (text)",    "as_stored"),
    ("M/DD/YY H:MM",        "m/dd/yy h:mm"),
    ("MM/DD/YYYY HH:MM:SS", "mm/dd/yyyy hh:mm:ss"),
    ("YYYY-MM-DD HH:MM",    "yyyy-mm-dd hh:mm"),
    ("D-MMM-YY H:MM AM/PM", "d-mmm-yy h:mm AM/PM"),
    ("D-MMM-YYYY",          "d-mmm-yyyy"),
    ("DD/MM/YYYY HH:MM",    "dd/mm/yyyy hh:mm"),
]


def _parse_timestamp(ts_string: str):
    """Try to parse *ts_string* using known Agilent patterns.
    Returns a :class:`datetime.datetime` on success, or ``None``."""
    for pattern in _TS_PARSE_FMTS:
        try:
            return datetime.datetime.strptime(ts_string.strip(), pattern)
        except ValueError:
            continue
    return None


def format_timestamp(ts_string: str, fmt: str) -> str:
    """Legacy helper: format *ts_string* as a plain string using a strftime *fmt*.
    Falls back to the original string on failure."""
    if fmt == 'as_stored' or not fmt:
        return ts_string
    dt = _parse_timestamp(ts_string)
    if dt is not None:
        return dt.strftime(fmt)
    return ts_string


def write_timestamp_cell(ws, row: int, col: int, ts_string: str, fmt: str, font):
    """Write a timestamp to *ws* at (*row*, *col*).

    * ``fmt == 'as_stored'``   → plain string, no number format.
    * ``fmt`` starts with '%'  → strftime text cell (backward-compat).
    * anything else            → parse to datetime, write as Excel native date
                                 with ``cell.number_format = fmt``.
    """
    cell = ws.cell(row=row, column=col)
    cell.font = font

    if fmt == 'as_stored' or not fmt:
        cell.value = ts_string
        return cell

    dt = _parse_timestamp(ts_string)

    if fmt.startswith('%'):
        # strftime string → text cell
        cell.value = dt.strftime(fmt) if dt else ts_string
    else:
        # Excel number-format string → native datetime cell
        if dt is not None:
            cell.value = dt
            cell.number_format = fmt
        else:
            cell.value = ts_string  # couldn't parse; write raw

    return cell


class ProcessingThread(QThread):
    """Thread for processing JSON files to Excel."""

    progress_update = Signal(str)  # Progress message
    finished = Signal(bool, str)   # Success, message

    def __init__(self, directory, output_file, format_config=None):
        super().__init__()
        self.directory = directory
        self.output_file = output_file
        self.format_config = format_config or copy.deepcopy(DEFAULT_FORMAT_CONFIG)

    def run(self):
        """Process JSON files to Excel in background thread."""
        try:
            self.progress_update.emit("Starting processing...")
            success = self.process_json_to_excel(self.directory, self.output_file)
            if success:
                self.finished.emit(True, f"Successfully created Excel file: {self.output_file}")
            else:
                self.finished.emit(False, "Processing completed but no JSON files were found.")
        except Exception as e:
            self.finished.emit(False, f"Error during processing: {str(e)}")

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _get_peak_value(self, peak: dict, key: str):
        """Return the cell value for *key* from a peak dict, with type coercion."""
        if key == 'quality_issues':
            # Derived: combine is_saturated / is_convoluted / quality_issues list
            issues = []
            if peak.get("is_saturated"):
                issues.append("Saturated")
            if peak.get("is_convoluted"):
                issues.append("Convoluted")
            raw = peak.get("quality_issues")
            if raw:
                if isinstance(raw, list):
                    issues.extend(raw)
                else:
                    issues.append(str(raw))
            return ", ".join(issues) if issues else ""

        _float_keys = {'retention_time', 'width', 'area', 'start_time', 'end_time',
                       'Qual', 'mol_C', 'mol_C_percent', 'mol', 'mass_mg',
                       'mol_percent', 'wt_percent'}
        _int_keys = {'peak_number', 'num_carbons'}

        raw = peak.get(key)
        if raw is None:
            return ""
        try:
            if key in _int_keys:
                return int(raw)
            if key in _float_keys:
                return round(float(raw), 4)
        except (ValueError, TypeError):
            pass
        return raw

    # ------------------------------------------------------------------
    # Main conversion
    # ------------------------------------------------------------------

    def process_json_to_excel(self, directory, output_file):
        """Process JSON files to Excel using self.format_config.

        Files that share the same directory (i.e. same .D folder) are written
        side-by-side on the same rows, separated by one blank column.  After
        all channels for a folder are written the next folder starts below the
        tallest channel block.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Chromatograms"

        header_font = Font(bold=True, size=12)
        subheader_font = Font(bold=True, size=10)
        quality_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        cfg = self.format_config
        enabled_headers = [h for h in cfg['header_fields'] if h['enabled']]
        enabled_columns = [c for c in cfg['columns'] if c['enabled']]
        ts_fmt = cfg.get('timestamp_format', 'as_stored')
        n_cols = len(enabled_columns)

        # Width of a single channel block in columns.
        # Header rows that show labels occupy 2 cols; peak table occupies n_cols.
        # We use whichever is wider so nothing overlaps.
        any_label = any(hf.get('show_label', True) for hf in enabled_headers)
        channel_width = max(n_cols, 2 if any_label else 1)

        starting_row = 1
        files_processed = 0

        self.progress_update.emit("Scanning for JSON files...")

        # Collect directories that contain JSON files, sorted chronologically
        # by the timestamp in the first JSON file of each directory.
        dir_file_pairs = []
        for root, _, files in os.walk(directory):
            json_files = sorted(f for f in files if f.endswith('.json'))
            if json_files:
                dir_file_pairs.append((root, json_files))

        def _dir_sort_key(pair):
            """Return a datetime for the first JSON file in the directory."""
            root, jfiles = pair
            try:
                with open(os.path.join(root, jfiles[0]), 'r', encoding='utf-8') as fh:
                    ts = json.load(fh).get('timestamp', '')
                dt = _parse_timestamp(str(ts)) if ts else None
                if dt is not None:
                    return dt
            except Exception:
                pass
            return datetime.datetime.max  # unparseable → sort to end

        dir_file_pairs.sort(key=_dir_sort_key)

        for root, json_files in dir_file_pairs:
            current_col = 1
            max_row_in_group = starting_row  # track deepest row across all channels

            for file in json_files:
                json_path = os.path.join(root, file)
                self.progress_update.emit(f"Processing {file}...")

                try:
                    with open(json_path, 'r', encoding='utf-8') as f:
                        json_data = json.load(f)

                    current_row = starting_row
                    col_offset = current_col

                    # --- Header fields ---
                    for hf in enabled_headers:
                        label = hf['custom_label'].strip() or hf['label']
                        raw_value = json_data.get(hf['key'], "")
                        if hf.get('show_label', True):
                            cell = ws.cell(row=current_row, column=col_offset, value=label)
                            cell.font = header_font
                            val_col = col_offset + 1
                        else:
                            val_col = col_offset

                        if hf['key'] == 'timestamp' and raw_value:
                            write_timestamp_cell(ws, current_row, val_col, str(raw_value), ts_fmt, header_font)
                        else:
                            cell = ws.cell(row=current_row, column=val_col, value=raw_value)
                            cell.font = header_font
                        current_row += 1

                    # --- Peak table column headers (no blank spacer row) ---
                    for i, col_cfg in enumerate(enabled_columns):
                        col_name = col_cfg['custom_name'].strip() or col_cfg['display']
                        cell = ws.cell(row=current_row, column=col_offset + i, value=col_name)
                        cell.font = subheader_font
                        cell.alignment = Alignment(horizontal='center')
                    current_row += 1

                    # --- Peak rows ---
                    for peak in json_data.get("peaks", []):
                        has_quality_issue = bool(self._get_peak_value(peak, 'quality_issues'))
                        for i, col_cfg in enumerate(enabled_columns):
                            value = self._get_peak_value(peak, col_cfg['key'])
                            ws.cell(row=current_row, column=col_offset + i, value=value)
                        if has_quality_issue:
                            for i in range(n_cols):
                                ws.cell(row=current_row, column=col_offset + i).fill = quality_fill
                        current_row += 1

                    max_row_in_group = max(max_row_in_group, current_row)
                    # Advance to next channel column (+1 for spacer)
                    current_col += channel_width + 1
                    files_processed += 1

                except Exception as e:
                    self.progress_update.emit(f"Error reading {json_path}: {e}")
                    print(f"Error reading {json_path}: {e}")

            # Move down past the tallest channel in this group
            starting_row = max_row_in_group + 2

        if files_processed > 0:
            wb.save(output_file)
            self.progress_update.emit(f"Successfully processed {files_processed} files")
            return True
        else:
            return False


class UpdateJsonThread(QThread):
    """Thread for updating JSON files from Excel compound IDs."""
    
    progress_update = Signal(str)  # Progress message
    finished = Signal(bool, str)   # Success, message
    
    def __init__(self, excel_file, json_directory):
        super().__init__()
        self.excel_file = excel_file
        self.json_directory = json_directory
    
    def run(self):
        """Update JSON files with compound IDs from Excel in background thread."""
        try:
            self.progress_update.emit("Reading Excel file...")
            success, files_updated = self.update_json_from_excel(self.excel_file, self.json_directory)
            
            if success:
                self.finished.emit(True, f"Successfully updated {files_updated} JSON files with compound IDs")
            else:
                self.finished.emit(False, "No JSON files were updated. Check file formats and data structure.")
                
        except Exception as e:
            self.finished.emit(False, f"Error during update: {str(e)}")

    def update_json_from_excel(self, excel_file, json_directory):
        """Update JSON files with compound IDs from Excel file."""
        try:
            # Load the Excel workbook
            wb = load_workbook(excel_file)
            ws = wb.active
            
            # Parse the Excel file to extract compound ID mappings
            compound_mappings = {}  # {sample_id: {peak_number: compound_id}}
            current_sample_id = None
            
            self.progress_update.emit("Parsing Excel data...")
            
            for row in ws.iter_rows(values_only=True):
                if not any(row):  # Skip empty rows
                    continue
                
                # Check if this row contains sample ID
                if row[0] == "Sample ID:":
                    current_sample_id = row[1] if len(row) > 1 else None
                    if current_sample_id:
                        compound_mappings[current_sample_id] = {}
                    continue
                
                # Check if this is a peaks data row (has Compound ID in first column)
                if (current_sample_id and len(row) >= 3 and 
                    isinstance(row[1], (int, float)) and  # Peak number
                    isinstance(row[2], (int, float))):    # Retention time
                    
                    compound_id = row[0] if row[0] not in [None, "Compound ID"] else "Unknown"
                    peak_number = int(row[1])
                    
                    compound_mappings[current_sample_id][peak_number] = compound_id
            
            if not compound_mappings:
                self.progress_update.emit("No compound ID mappings found in Excel file")
                return False, 0
            
            self.progress_update.emit(f"Found mappings for {len(compound_mappings)} samples")
            
            # Update JSON files
            files_updated = 0
            
            for root, _, files in os.walk(json_directory):
                json_files = [f for f in files if f.endswith('.json')]
                
                for file in json_files:
                    json_path = os.path.join(root, file)
                    
                    try:
                        with open(json_path, 'r', encoding='utf-8') as f:
                            json_data = json.load(f)
                        
                        sample_id = json_data.get("sample_id")
                        if not sample_id or sample_id not in compound_mappings:
                            continue
                        
                        self.progress_update.emit(f"Updating {file}...")
                        
                        # Update compound IDs for each peak
                        updated = False
                        sample_mappings = compound_mappings[sample_id]
                        
                        for peak in json_data.get("peaks", []):
                            peak_number = peak.get("peak_number")
                            if peak_number in sample_mappings:
                                old_compound_id = peak.get("Compound ID", "Unknown")
                                new_compound_id = sample_mappings[peak_number]
                                
                                if old_compound_id != new_compound_id:
                                    peak["Compound ID"] = new_compound_id
                                    updated = True
                        
                        # Save the updated JSON file
                        if updated:
                            with open(json_path, 'w', encoding='utf-8') as f:
                                json.dump(json_data, f, indent=2, ensure_ascii=False)
                            files_updated += 1
                    
                    except Exception as e:
                        self.progress_update.emit(f"Error updating {json_path}: {e}")
                        continue
            
            return files_updated > 0, files_updated
            
        except Exception as e:
            self.progress_update.emit(f"Error reading Excel file: {e}")
            return False, 0


class JsonToExcelConverter(QDialog):
    """Dialog for JSON to Excel conversion."""
    
    def __init__(self, parent=None, initial_directory=None):
        super().__init__(parent)
        self.setWindowTitle("ChromaKit-MS: JSON ↔ Excel Converter")
        self.setModal(True)
        self.resize(750, 700)

        # QSettings for format persistence
        self.settings = QSettings("CalebCoatney", "ChromaKit")

        # Initialize variables for JSON to Excel
        self.input_directory = initial_directory or ""
        self.output_file = ""

        # Initialize variables for Excel to JSON
        self.excel_input_file = ""
        self.json_directory = initial_directory or ""

        self.init_ui()

        # Load last-used format config
        self._load_last_format_config()

        # Pre-populate directory labels if initial_directory was provided
        if initial_directory:
            self.input_label.setText(initial_directory)
            self.json_dir_label.setText(initial_directory)
            self.log_text.append(f"Pre-loaded directory: {initial_directory}")
            self.update_process_button()
            self.update_json_update_button()

    def init_ui(self):
        """Initialize the user interface."""
        layout = QVBoxLayout(self)

        # Title
        title_label = QLabel("ChromaKit-MS JSON ↔ Excel Converter")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title_label.setFont(title_font)
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)

        # Description
        desc_label = QLabel(
            "Convert between ChromaKit-MS JSON files and Excel format, "
            "including compound ID management."
        )
        desc_label.setWordWrap(True)
        desc_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(desc_label)

        layout.addSpacing(10)

        # Create tabbed interface
        self.tab_widget = QTabWidget()

        # JSON to Excel tab
        self.json_to_excel_tab = self.create_json_to_excel_tab()
        self.tab_widget.addTab(self.json_to_excel_tab, "JSON → Excel")

        # Format tab
        self.format_tab = self.create_format_tab()
        self.tab_widget.addTab(self.format_tab, "Format")

        # Excel to JSON tab
        self.excel_to_json_tab = self.create_excel_to_json_tab()
        self.tab_widget.addTab(self.excel_to_json_tab, "Excel → JSON")

        layout.addWidget(self.tab_widget)

        # Shared progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)

        # Shared log output
        log_group = QGroupBox("Processing Log")
        log_layout = QVBoxLayout(log_group)

        self.log_text = QTextEdit()
        self.log_text.setMaximumHeight(150)
        self.log_text.setReadOnly(True)
        log_layout.addWidget(self.log_text)

        layout.addWidget(log_group)
    
    def create_json_to_excel_tab(self):
        """Create the JSON to Excel conversion tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Input directory selection
        input_group = QGroupBox("Input Directory")
        input_layout = QVBoxLayout(input_group)
        
        input_button_layout = QHBoxLayout()
        self.input_label = QLabel("No directory selected")
        self.input_label.setStyleSheet("QLabel { background-color: #f0f0f0; padding: 5px; border: 1px solid #ccc; }")
        self.select_input_btn = QPushButton("Select Directory")
        self.select_input_btn.clicked.connect(self.select_input_directory)
        
        input_button_layout.addWidget(self.input_label, 1)
        input_button_layout.addWidget(self.select_input_btn)
        input_layout.addLayout(input_button_layout)
        
        layout.addWidget(input_group)
        
        # Output file selection
        output_group = QGroupBox("Output Excel File")
        output_layout = QVBoxLayout(output_group)
        
        output_button_layout = QHBoxLayout()
        self.output_label = QLabel("No output file selected")
        self.output_label.setStyleSheet("QLabel { background-color: #f0f0f0; padding: 5px; border: 1px solid #ccc; }")
        self.select_output_btn = QPushButton("Select Output File")
        self.select_output_btn.clicked.connect(self.select_output_file)
        
        output_button_layout.addWidget(self.output_label, 1)
        output_button_layout.addWidget(self.select_output_btn)
        output_layout.addLayout(output_button_layout)
        
        layout.addWidget(output_group)
        
        # Process button
        self.process_btn = QPushButton("Convert to Excel")
        self.process_btn.setStyleSheet("QPushButton { background-color: #4CAF50; color: white; font-weight: bold; padding: 10px; }")
        self.process_btn.clicked.connect(self.start_processing)
        self.process_btn.setEnabled(False)
        layout.addWidget(self.process_btn)

        layout.addStretch()
        return tab
    
    def create_excel_to_json_tab(self):
        """Create the Excel to JSON update tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Description
        desc_label = QLabel(
            "Update JSON files with compound IDs from an Excel file that was "
            "previously exported and manually annotated with compound identifications."
        )
        desc_label.setWordWrap(True)
        layout.addWidget(desc_label)
        
        layout.addSpacing(10)
        
        # Excel input file selection
        excel_input_group = QGroupBox("Input Excel File")
        excel_input_layout = QVBoxLayout(excel_input_group)
        
        excel_input_button_layout = QHBoxLayout()
        self.excel_input_label = QLabel("No Excel file selected")
        self.excel_input_label.setStyleSheet("QLabel { background-color: #f0f0f0; padding: 5px; border: 1px solid #ccc; }")
        self.select_excel_input_btn = QPushButton("Select Excel File")
        self.select_excel_input_btn.clicked.connect(self.select_excel_input_file)
        
        excel_input_button_layout.addWidget(self.excel_input_label, 1)
        excel_input_button_layout.addWidget(self.select_excel_input_btn)
        excel_input_layout.addLayout(excel_input_button_layout)
        
        layout.addWidget(excel_input_group)
        
        # JSON directory selection
        json_dir_group = QGroupBox("JSON Files Directory")
        json_dir_layout = QVBoxLayout(json_dir_group)
        
        json_dir_button_layout = QHBoxLayout()
        self.json_dir_label = QLabel("No directory selected")
        self.json_dir_label.setStyleSheet("QLabel { background-color: #f0f0f0; padding: 5px; border: 1px solid #ccc; }")
        self.select_json_dir_btn = QPushButton("Select Directory")
        self.select_json_dir_btn.clicked.connect(self.select_json_directory)
        
        json_dir_button_layout.addWidget(self.json_dir_label, 1)
        json_dir_button_layout.addWidget(self.select_json_dir_btn)
        json_dir_layout.addLayout(json_dir_button_layout)
        
        layout.addWidget(json_dir_group)
        
        # Update button
        self.update_json_btn = QPushButton("Update JSON Files")
        self.update_json_btn.setStyleSheet("QPushButton { background-color: #2196F3; color: white; font-weight: bold; padding: 10px; }")
        self.update_json_btn.clicked.connect(self.start_json_update)
        self.update_json_btn.setEnabled(False)
        layout.addWidget(self.update_json_btn)
        
        layout.addStretch()
        return tab
    
    def select_input_directory(self):
        """Select input directory containing .D folders with JSON files."""
        directory = QFileDialog.getExistingDirectory(
            self, 
            "Select Directory Containing .D Folders",
            ""
        )
        
        if directory:
            self.input_directory = directory
            self.input_label.setText(directory)
            self.log_text.append(f"Selected input directory: {directory}")
            self.update_process_button()
    
    def select_output_file(self):
        """Select output Excel file location."""
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel File",
            "ChromaKit_Results.xlsx",
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if file_path:
            if not file_path.endswith('.xlsx'):
                file_path += '.xlsx'
            self.output_file = file_path
            self.output_label.setText(file_path)
            self.log_text.append(f"Selected output file: {file_path}")
            self.update_process_button()
    
    def update_process_button(self):
        """Enable process button if both input and output are selected."""
        self.process_btn.setEnabled(bool(self.input_directory and self.output_file))
    
    def start_processing(self):
        """Start the JSON to Excel conversion process."""
        if not self.input_directory or not self.output_file:
            QMessageBox.warning(self, "Error", "Please select both input directory and output file.")
            return

        # Persist current format config before processing
        self._save_last_format_config()

        # Disable UI during processing
        self.process_btn.setEnabled(False)
        self.select_input_btn.setEnabled(False)
        self.select_output_btn.setEnabled(False)

        # Show progress bar
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 0)  # Indeterminate progress

        # Clear log
        self.log_text.clear()
        self.log_text.append("Starting conversion process...")

        # Start processing thread
        self.processing_thread = ProcessingThread(
            self.input_directory,
            self.output_file,
            self.get_current_format_config()
        )
        self.processing_thread.progress_update.connect(self.update_progress)
        self.processing_thread.finished.connect(self.processing_finished)
        self.processing_thread.start()

    def update_progress(self, message):
        """Update progress log."""
        self.log_text.append(message)
        self.log_text.ensureCursorVisible()
        QApplication.processEvents()

    def processing_finished(self, success, message):
        """Handle processing completion."""
        # Re-enable UI
        self.process_btn.setEnabled(True)
        self.select_input_btn.setEnabled(True)
        self.select_output_btn.setEnabled(True)

        # Hide progress bar
        self.progress_bar.setVisible(False)

        # Show completion message
        self.log_text.append(f"Processing completed: {message}")

        if success:
            QMessageBox.information(self, "Success", message)
        else:
            QMessageBox.warning(self, "Warning", message)

    # ------------------------------------------------------------------
    # Format tab
    # ------------------------------------------------------------------

    def create_format_tab(self):
        """Create the Format Settings tab."""
        tab = QWidget()
        outer_layout = QVBoxLayout(tab)

        # --- Timestamp format ---
        ts_group = QGroupBox("Timestamp Format")
        ts_layout = QHBoxLayout(ts_group)
        ts_layout.addWidget(QLabel("Output format:"))
        self.ts_format_combo = QComboBox()
        for label, fmt in TIMESTAMP_FORMAT_OPTIONS:
            self.ts_format_combo.addItem(label, fmt)
        self.ts_format_combo.setToolTip(
            "All formats except 'As stored' write a native Excel date/time cell.\n"
            "'As stored' writes the raw timestamp string as text."
        )
        ts_layout.addWidget(self.ts_format_combo, 1)
        outer_layout.addWidget(ts_group)

        # --- Header fields (drag-to-reorder) ---
        header_group = QGroupBox(
            "Header Rows  —  drag to reorder   ☑ include   rename: optional label   show label: 'Label: value' vs just 'value'"
        )
        header_outer = QVBoxLayout(header_group)

        hint = QLabel("Drag rows to reorder")
        hint.setStyleSheet("color: gray; font-style: italic;")
        header_outer.addWidget(hint)

        self._header_list = QListWidget()
        self._header_list.setDragDropMode(QAbstractItemView.InternalMove)
        self._header_list.setDefaultDropAction(Qt.MoveAction)
        self._header_list.setSelectionMode(QAbstractItemView.SingleSelection)
        self._header_list.setFixedHeight(6 * 38)  # 6 rows × row height
        # After a drag-drop reorder the item data (UserRole) stays with its item, but
        # the item widgets are visually stale – rebuild them from the data.
        self._header_list.model().rowsMoved.connect(lambda *_: self._rebuild_header_item_widgets())

        for hf in copy.deepcopy(DEFAULT_FORMAT_CONFIG['header_fields']):
            item = QListWidgetItem()
            item.setData(Qt.UserRole, hf)
            item.setSizeHint(QSize(0, 36))
            self._header_list.addItem(item)
            self._header_list.setItemWidget(item, self._make_header_item_widget(item))

        header_outer.addWidget(self._header_list)
        outer_layout.addWidget(header_group)

        # --- Peak table columns ---
        col_group = QGroupBox("Peak Table Columns  (uncheck to omit; optionally rename)")
        col_outer = QVBoxLayout(col_group)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setMinimumHeight(220)
        col_container = QWidget()
        col_layout = QFormLayout(col_container)
        col_layout.setLabelAlignment(Qt.AlignLeft)

        self._column_widgets = []
        for col_cfg in DEFAULT_FORMAT_CONFIG['columns']:
            cb = QCheckBox(col_cfg['display'])
            cb.setChecked(col_cfg['enabled'])
            rename = QLineEdit()
            rename.setPlaceholderText(col_cfg['display'])
            rename.setFixedWidth(180)
            rename.setEnabled(col_cfg['enabled'])
            cb.toggled.connect(rename.setEnabled)
            col_layout.addRow(cb, rename)
            self._column_widgets.append((cb, rename))

        scroll.setWidget(col_container)
        col_outer.addWidget(scroll)
        outer_layout.addWidget(col_group)

        # --- Presets ---
        preset_group = QGroupBox("Presets")
        preset_layout = QHBoxLayout(preset_group)
        preset_layout.addWidget(QLabel("Preset:"))
        self.preset_combo = QComboBox()
        self.preset_combo.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        preset_layout.addWidget(self.preset_combo, 1)
        load_btn = QPushButton("Load")
        load_btn.clicked.connect(self._load_selected_preset)
        preset_layout.addWidget(load_btn)
        save_btn = QPushButton("Save As…")
        save_btn.clicked.connect(self._save_preset_as)
        preset_layout.addWidget(save_btn)
        delete_btn = QPushButton("Delete")
        delete_btn.clicked.connect(self._delete_selected_preset)
        preset_layout.addWidget(delete_btn)
        outer_layout.addWidget(preset_group)

        self._refresh_preset_combo()
        return tab

    # ------------------------------------------------------------------
    # Drag-drop header list helpers
    # ------------------------------------------------------------------

    def _make_header_item_widget(self, item: QListWidgetItem) -> QWidget:
        """Create the inline widget for one header-list item."""
        hf = item.data(Qt.UserRole)

        row_w = QWidget()
        row_l = QHBoxLayout(row_w)
        row_l.setContentsMargins(4, 2, 4, 2)

        cb = QCheckBox(hf['label'])
        cb.setChecked(hf.get('enabled', True))
        cb.toggled.connect(lambda checked, it=item: self._update_header_item_data(it, 'enabled', checked))
        row_l.addWidget(cb)

        rename = QLineEdit()
        rename.setPlaceholderText(hf['label'])
        rename.setText(hf.get('custom_label', ''))
        rename.setFixedWidth(140)
        rename.setEnabled(hf.get('enabled', True))
        cb.toggled.connect(rename.setEnabled)
        rename.textChanged.connect(
            lambda text, it=item: self._update_header_item_data(it, 'custom_label', text.strip())
        )
        row_l.addWidget(rename)

        show_label_cb = QCheckBox("show label")
        show_label_cb.setChecked(hf.get('show_label', True))
        show_label_cb.setToolTip(
            "When checked: writes the field label (e.g. 'Timestamp:') in the first cell "
            "and the value in the next cell.\n"
            "When unchecked: writes only the value in the first cell."
        )
        show_label_cb.toggled.connect(
            lambda checked, it=item: self._update_header_item_data(it, 'show_label', checked)
        )
        row_l.addWidget(show_label_cb)
        row_l.addStretch()

        return row_w

    def _update_header_item_data(self, item: QListWidgetItem, key: str, value):
        """Keep the item's UserRole dict in sync with widget state."""
        data = item.data(Qt.UserRole) or {}
        data[key] = value
        item.setData(Qt.UserRole, data)

    def _rebuild_header_item_widgets(self):
        """Re-attach item widgets after a drag-drop reorder (widgets don't move with items)."""
        for i in range(self._header_list.count()):
            item = self._header_list.item(i)
            item.setSizeHint(QSize(0, 36))
            self._header_list.setItemWidget(item, self._make_header_item_widget(item))

    def _collect_header_state(self) -> list:
        """Return current header fields list (order + state) from the list widget."""
        return [
            self._header_list.item(i).data(Qt.UserRole)
            for i in range(self._header_list.count())
        ]

    def _populate_header_list(self, fields: list):
        """Replace all items in the header list with *fields*."""
        self._header_list.clear()
        for hf in fields:
            item = QListWidgetItem()
            item.setData(Qt.UserRole, copy.deepcopy(hf))
            item.setSizeHint(QSize(0, 36))
            self._header_list.addItem(item)
            self._header_list.setItemWidget(item, self._make_header_item_widget(item))

    # ------------------------------------------------------------------
    # Format config get/set helpers
    # ------------------------------------------------------------------

    def get_current_format_config(self) -> dict:
        """Build a format_config dict from the current Format tab UI state."""
        cfg = {
            'timestamp_format': self.ts_format_combo.currentData(),
            'header_fields': self._collect_header_state(),
            'columns': [],
        }

        for i, col in enumerate(DEFAULT_FORMAT_CONFIG['columns']):
            cb, rename = self._column_widgets[i]
            cfg['columns'].append({
                'key': col['key'],
                'display': col['display'],
                'enabled': cb.isChecked(),
                'custom_name': rename.text().strip(),
            })

        return cfg

    def _apply_format_config(self, cfg: dict):
        """Push a format_config dict into the Format tab UI widgets."""
        # Timestamp
        ts_fmt = cfg.get('timestamp_format', 'as_stored')
        for i in range(self.ts_format_combo.count()):
            if self.ts_format_combo.itemData(i) == ts_fmt:
                self.ts_format_combo.setCurrentIndex(i)
                break

        # Header fields: restore saved order; append any missing keys at the end.
        default_by_key = {hf['key']: hf for hf in DEFAULT_FORMAT_CONFIG['header_fields']}
        saved_fields = cfg.get('header_fields', [])
        saved_keys = [hf['key'] for hf in saved_fields]

        restored = []
        for hf in saved_fields:
            default = default_by_key.get(hf['key'], hf)
            restored.append({
                'key': hf['key'],
                'label': default['label'],
                'enabled': hf.get('enabled', True),
                'custom_label': hf.get('custom_label', ''),
                'show_label': hf.get('show_label', True),
            })
        for key, default in default_by_key.items():
            if key not in saved_keys:
                restored.append(copy.deepcopy(default))

        self._populate_header_list(restored)

        # Columns – match by key
        col_by_key = {c['key']: c for c in cfg.get('columns', [])}
        for i, default_col in enumerate(DEFAULT_FORMAT_CONFIG['columns']):
            cb, rename = self._column_widgets[i]
            col = col_by_key.get(default_col['key'], default_col)
            cb.setChecked(col.get('enabled', default_col['enabled']))
            rename.setText(col.get('custom_name', ''))

    # ------------------------------------------------------------------
    # Preset management
    # ------------------------------------------------------------------

    def _get_all_presets(self) -> dict:
        raw = self.settings.value("xlsx_format_presets", "{}")
        try:
            return json.loads(raw)
        except Exception:
            return {}

    def _save_all_presets(self, presets: dict):
        self.settings.setValue("xlsx_format_presets", json.dumps(presets))

    def _refresh_preset_combo(self):
        self.preset_combo.clear()
        for name in sorted(self._get_all_presets().keys()):
            self.preset_combo.addItem(name)

    def _save_preset_as(self):
        name, ok = QInputDialog.getText(self, "Save Preset", "Preset name:")
        if not ok or not name.strip():
            return
        name = name.strip()
        presets = self._get_all_presets()
        presets[name] = self.get_current_format_config()
        self._save_all_presets(presets)
        self._refresh_preset_combo()
        # Select the just-saved preset
        idx = self.preset_combo.findText(name)
        if idx >= 0:
            self.preset_combo.setCurrentIndex(idx)
        self.log_text.append(f"Preset '{name}' saved.")

    def _load_selected_preset(self):
        name = self.preset_combo.currentText()
        if not name:
            return
        presets = self._get_all_presets()
        if name not in presets:
            QMessageBox.warning(self, "Preset not found", f"Could not find preset '{name}'.")
            return
        self._apply_format_config(presets[name])
        self.log_text.append(f"Preset '{name}' loaded.")

    def _delete_selected_preset(self):
        name = self.preset_combo.currentText()
        if not name:
            return
        reply = QMessageBox.question(
            self, "Delete Preset",
            f"Delete preset '{name}'?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            presets = self._get_all_presets()
            presets.pop(name, None)
            self._save_all_presets(presets)
            self._refresh_preset_combo()
            self.log_text.append(f"Preset '{name}' deleted.")

    def _save_last_format_config(self):
        self.settings.setValue("xlsx_last_format", json.dumps(self.get_current_format_config()))

    def _load_last_format_config(self):
        raw = self.settings.value("xlsx_last_format", "")
        if raw:
            try:
                self._apply_format_config(json.loads(raw))
            except Exception:
                pass

    def select_excel_input_file(self):
        """Select Excel input file for updating JSON files."""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Excel File",
            "",
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if file_path:
            self.excel_input_file = file_path
            self.excel_input_label.setText(file_path)
            self.log_text.append(f"Selected Excel input file: {file_path}")
            self.update_json_update_button()
    
    def select_json_directory(self):
        """Select directory containing JSON files to update."""
        directory = QFileDialog.getExistingDirectory(
            self, 
            "Select Directory Containing JSON Files",
            ""
        )
        
        if directory:
            self.json_directory = directory
            self.json_dir_label.setText(directory)
            self.log_text.append(f"Selected JSON directory: {directory}")
            self.update_json_update_button()
    
    def update_json_update_button(self):
        """Enable JSON update button if both Excel file and JSON directory are selected."""
        self.update_json_btn.setEnabled(bool(self.excel_input_file and self.json_directory))
    
    def start_json_update(self):
        """Start the Excel to JSON update process."""
        if not self.excel_input_file or not self.json_directory:
            QMessageBox.warning(self, "Error", "Please select both Excel file and JSON directory.")
            return
        
        # Disable UI during processing
        self.update_json_btn.setEnabled(False)
        self.select_excel_input_btn.setEnabled(False)
        self.select_json_dir_btn.setEnabled(False)
        
        # Show progress bar
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 0)  # Indeterminate progress
        
        # Clear log
        self.log_text.clear()
        self.log_text.append("Starting JSON update process...")
        
        # Start update thread
        self.update_thread = UpdateJsonThread(
            self.excel_input_file,
            self.json_directory
        )
        self.update_thread.progress_update.connect(self.update_progress)
        self.update_thread.finished.connect(self.json_update_finished)
        self.update_thread.start()
    
    def json_update_finished(self, success, message):
        """Handle JSON update completion."""
        # Re-enable UI
        self.update_json_btn.setEnabled(True)
        self.select_excel_input_btn.setEnabled(True)
        self.select_json_dir_btn.setEnabled(True)
        
        # Hide progress bar
        self.progress_bar.setVisible(False)
        
        # Show completion message
        self.log_text.append(f"JSON update completed: {message}")
        
        if success:
            QMessageBox.information(self, "Success", message)
        else:
            QMessageBox.warning(self, "Warning", message)


def main():
    """Main entry point."""
    app = QApplication(sys.argv)
    
    # Set application properties
    app.setApplicationName("ChromaKit-MS JSON ↔ Excel Converter")
    app.setApplicationVersion("2.0.0")
    
    # Create and show main window
    window = JsonToExcelConverter()
    window.show()
    
    # Run the application
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
