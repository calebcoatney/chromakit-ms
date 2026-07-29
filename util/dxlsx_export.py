# -*- coding: utf-8 -*-
"""Standalone utility: export Agilent .D directories to a single .xlsx workbook.

One worksheet per .D folder, a shared Time (min) column, and one column per
selected signal (GC .ch detectors and/or MS data.ms TIC), interpolated onto a
common retention-time grid.

Run: python util/dxlsx_export.py
"""

import os
import sys
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font


# ---------------------------------------------------------------------------
# Data-reading helpers (rainbow + numpy only)
# ---------------------------------------------------------------------------

def list_signals(data_dir):
    """Return signal filenames in a rainbow data dir: '*.ch' plus 'data.ms'."""
    names = [str(f.name) for f in data_dir.datafiles]
    signals = [n for n in names if n.endswith(".ch")]
    if "data.ms" in names:
        signals.append("data.ms")
    return signals


def read_signal(data_dir, signal):
    """Return (x_minutes, y) for a signal. MS 'data.ms' returns the TIC."""
    f = data_dir.get_file(signal)
    x = np.asarray(f.xlabels, dtype=float).flatten()
    if signal == "data.ms":
        y = np.sum(np.asarray(f.data, dtype=float), axis=1)
    else:
        y = np.asarray(f.data, dtype=float).flatten()
    return x, y


def read_notebook(data_dir, d_path):
    """Best-effort notebook name: detector-file metadata -> dir metadata ->
    data_dir.name -> folder basename."""
    detector_files = [f for f in data_dir.datafiles if str(f.name).endswith(".ch")]
    if detector_files:
        nb = detector_files[0].metadata.get("notebook")
        if nb:
            return str(nb)
    nb = data_dir.metadata.get("notebook")
    if nb:
        return str(nb)
    name = getattr(data_dir, "name", None)
    if name:
        return str(name)
    return os.path.splitext(os.path.basename(os.path.normpath(d_path)))[0]


def build_time_grid(sig_x, skip_solvent_delay, has_ms, ms_x, n):
    """Build a common time grid (minutes) over the union range of all signals.

    sig_x: dict signal_name -> native x array.
    If skip_solvent_delay and has_ms, clip the start to min(ms_x).
    """
    starts = [np.min(x) for x in sig_x.values() if len(x)]
    ends = [np.max(x) for x in sig_x.values() if len(x)]
    if not starts:
        raise ValueError("build_time_grid: no non-empty signals provided")
    t_min = min(starts)
    t_max = max(ends)
    if skip_solvent_delay and has_ms and ms_x is not None and len(ms_x):
        t_min = float(np.min(ms_x))
    return np.linspace(t_min, t_max, int(n))


def resample_to_grid(grid, x, y):
    """np.interp onto grid, masking points outside [x.min(), x.max()] to NaN."""
    out = np.interp(grid, x, y)
    out = np.where((grid < np.min(x)) | (grid > np.max(x)), np.nan, out)
    return out


_INVALID_SHEET_CHARS = r'[]:*?/\\'


def safe_sheet_name(raw, used_names):
    """Sanitize to a valid, unique Excel sheet name (<=31 chars, no invalid
    chars, no leading/trailing spaces). Mutates used_names in place, adding the
    chosen name, and returns it."""
    name = "".join(c for c in str(raw) if c not in _INVALID_SHEET_CHARS).strip()
    name = name[:31].strip()
    if not name:
        name = "Sheet"
    base = name
    counter = 2
    while name in used_names:
        suffix = "_" + str(counter)
        name = base[:31 - len(suffix)].strip() + suffix
        counter += 1
    used_names.add(name)
    return name


def build_sheet_rows(data_dir, selected, skip_solvent_delay, n):
    """Return (header, rows) for one .D folder.

    header: ["Time (min)", <signal>, ...] for selected signals present here.
    rows: list of [time, val, ...] with NaN converted to None.
    """
    present = [s for s in selected if s in list_signals(data_dir)]
    if not present:
        return ["Time (min)"], []
    sig_xy = {s: read_signal(data_dir, s) for s in present}
    sig_x = {s: xy[0] for s, xy in sig_xy.items()}

    has_ms = "data.ms" in present
    ms_x = sig_x.get("data.ms")
    grid = build_time_grid(sig_x, skip_solvent_delay, has_ms, ms_x, n)

    columns = [grid]
    header = ["Time (min)"]
    for s in present:
        x, y = sig_xy[s]
        columns.append(resample_to_grid(grid, x, y))
        header.append(s)

    rows = []
    for i in range(len(grid)):
        row = []
        for col in columns:
            v = col[i]
            row.append(None if (isinstance(v, float) and np.isnan(v)) else float(v))
        rows.append(row)
    return header, rows


def write_workbook(out_path, sheets):
    """sheets: list of (sheet_name, header, rows). Writes bold header row."""
    if not sheets:
        raise ValueError("write_workbook: no sheets to write")
    wb = Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)
    for sheet_name, header, rows in sheets:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(header)
        for c in range(1, len(header) + 1):
            ws.cell(row=1, column=c).font = bold
        for row in rows:
            ws.append(row)
    wb.save(out_path)


def _default_reader(path):
    import rainbow as rb
    return rb.read(path)


def export_folders(folders, selected, skip_solvent_delay, n, out_path,
                   reader=_default_reader, log=print, progress=None):
    """Read each folder, build a sheet, write the workbook.

    Returns {'exported': int, 'skipped': int}. Failures are logged and skipped.
    """
    sheets = []
    used_names = set()
    exported = 0
    skipped = 0
    total = len(folders)
    for idx, folder in enumerate(folders):
        try:
            data_dir = reader(folder)
            header, rows = build_sheet_rows(data_dir, selected,
                                            skip_solvent_delay, n)
            if len(header) <= 1:
                log("Skipped (no selected signals present): " + str(folder))
                skipped += 1
            else:
                raw_name = read_notebook(data_dir, folder)
                sheet_name = safe_sheet_name(raw_name, used_names)
                sheets.append((sheet_name, header, rows))
                exported += 1
                log("Exported: " + str(folder) + " -> " + sheet_name)
        except Exception as e:  # noqa: BLE001 - report and continue
            log("Skipped (error): " + str(folder) + " (" + type(e).__name__ + ": " + str(e) + ")")
            skipped += 1
        if progress is not None:
            progress(int((idx + 1) / max(total, 1) * 100))

    if sheets:
        write_workbook(out_path, sheets)
    else:
        log("No sheets to write.")
    return {"exported": exported, "skipped": skipped}


# ---------------------------------------------------------------------------
# Task 7: QThread export worker
# ---------------------------------------------------------------------------

from PySide6.QtCore import QThread, Signal


class ExportWorker(QThread):
    progress = Signal(int)
    log = Signal(str)
    done = Signal(dict)

    def __init__(self, folders, selected, skip_solvent_delay, n, out_path):
        super().__init__()
        self._folders = folders
        self._selected = selected
        self._skip = skip_solvent_delay
        self._n = n
        self._out = out_path

    def run(self):
        try:
            result = export_folders(
                self._folders, self._selected, self._skip, self._n, self._out,
                log=self.log.emit, progress=self.progress.emit)
        except Exception as e:  # noqa: BLE001
            self.log.emit("Fatal error: " + str(e))
            result = {"exported": 0, "skipped": len(self._folders), "error": str(e)}
        self.done.emit(result)


# ---------------------------------------------------------------------------
# Task 8: GUI dialog
# ---------------------------------------------------------------------------

import rainbow as rb
from PySide6.QtWidgets import (
    QApplication, QDialog, QVBoxLayout, QHBoxLayout, QWidget, QPushButton,
    QLabel, QFileDialog, QTextEdit, QProgressBar, QMessageBox, QGroupBox,
    QCheckBox, QLineEdit, QScrollArea, QSpinBox, QListWidget,
    QAbstractItemView,
)
from PySide6.QtCore import Qt, QSettings


class ExportDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Agilent .D → xlsx Export")
        self.resize(640, 640)
        self._settings = QSettings("CalebCoatney", "ChromaKit")
        self._folders = []          # list[str]
        self._signal_checks = {}    # signal name -> QCheckBox
        self._worker = None
        self._build_ui()

    # -- UI construction -------------------------------------------------
    def _build_ui(self):
        layout = QVBoxLayout(self)

        # Folders
        fbox = QGroupBox("Agilent .D folders")
        fl = QVBoxLayout(fbox)
        self._folder_list = QListWidget()
        self._folder_list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        fl.addWidget(self._folder_list)
        frow = QHBoxLayout()
        add_btn = QPushButton("Add .D Folder(s)…")
        add_btn.clicked.connect(self._add_folders)
        rm_btn = QPushButton("Remove Selected")
        rm_btn.clicked.connect(self._remove_selected)
        frow.addWidget(add_btn)
        frow.addWidget(rm_btn)
        fl.addLayout(frow)
        layout.addWidget(fbox)

        # Signals
        sbox = QGroupBox("Signals to export")
        sl = QVBoxLayout(sbox)
        self._signal_area = QScrollArea()
        self._signal_area.setWidgetResizable(True)
        self._signal_inner = QWidget()
        self._signal_layout = QVBoxLayout(self._signal_inner)
        self._signal_area.setWidget(self._signal_inner)
        sl.addWidget(self._signal_area)
        layout.addWidget(sbox)

        # Options
        obox = QGroupBox("Options")
        ol = QHBoxLayout(obox)
        self._skip_check = QCheckBox("Skip solvent delay (clip to MS start when MS present)")
        self._skip_check.setChecked(True)
        ol.addWidget(self._skip_check)
        ol.addWidget(QLabel("Points per signal:"))
        self._points_spin = QSpinBox()
        self._points_spin.setRange(100, 200000)
        self._points_spin.setValue(int(self._settings.value("dxlsx/points", 10000)))
        ol.addWidget(self._points_spin)
        layout.addWidget(obox)

        # Output
        orow = QHBoxLayout()
        orow.addWidget(QLabel("Output:"))
        self._out_edit = QLineEdit()
        self._out_edit.setReadOnly(True)
        self._out_edit.setText(self._settings.value("dxlsx/out", ""))
        browse = QPushButton("Browse…")
        browse.clicked.connect(self._pick_output)
        orow.addWidget(self._out_edit)
        orow.addWidget(browse)
        layout.addLayout(orow)

        # Export + progress + log
        self._export_btn = QPushButton("Export")
        self._export_btn.clicked.connect(self._start_export)
        layout.addWidget(self._export_btn)
        self._progress = QProgressBar()
        layout.addWidget(self._progress)
        self._log = QTextEdit()
        self._log.setReadOnly(True)
        layout.addWidget(self._log)

    # -- Folder handling -------------------------------------------------
    def _add_folders(self):
        d = QFileDialog.getExistingDirectory(self, "Select an Agilent .D folder")
        if d and d not in self._folders:
            self._folders.append(d)
            self._folder_list.addItem(d)
            self._rescan_signals()

    def _remove_selected(self):
        for item in self._folder_list.selectedItems():
            path = item.text()
            if path in self._folders:
                self._folders.remove(path)
            self._folder_list.takeItem(self._folder_list.row(item))
        self._rescan_signals()

    def _rescan_signals(self):
        prev = {s: cb.isChecked() for s, cb in self._signal_checks.items()}
        union = []
        for folder in self._folders:
            try:
                data_dir = rb.read(folder)
                for s in list_signals(data_dir):
                    if s not in union:
                        union.append(s)
            except Exception as e:  # noqa: BLE001
                self._log.append("Could not scan " + folder + ": " + str(e))
        while self._signal_layout.count():
            item = self._signal_layout.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()
        self._signal_checks = {}
        for s in union:
            cb = QCheckBox(s)
            cb.setChecked(prev.get(s, True))
            self._signal_layout.addWidget(cb)
            self._signal_checks[s] = cb

    # -- Output ----------------------------------------------------------
    def _pick_output(self):
        start_dir = self._settings.value("dxlsx/out", "")
        path, _ = QFileDialog.getSaveFileName(
            self, "Save workbook", start_dir, "Excel Workbook (*.xlsx)")
        if path:
            if not path.lower().endswith(".xlsx"):
                path += ".xlsx"
            self._out_edit.setText(path)

    # -- Export ----------------------------------------------------------
    def _start_export(self):
        selected = [s for s, cb in self._signal_checks.items() if cb.isChecked()]
        out_path = self._out_edit.text().strip()
        if not self._folders:
            QMessageBox.warning(self, "No folders", "Add at least one .D folder.")
            return
        if not selected:
            QMessageBox.warning(self, "No signals", "Select at least one signal.")
            return
        if not out_path:
            QMessageBox.warning(self, "No output", "Choose an output .xlsx path.")
            return

        self._settings.setValue("dxlsx/points", self._points_spin.value())
        self._settings.setValue("dxlsx/out", out_path)

        self._export_btn.setEnabled(False)
        self._progress.setValue(0)
        self._log.clear()

        self._worker = ExportWorker(
            list(self._folders), selected, self._skip_check.isChecked(),
            self._points_spin.value(), out_path)
        self._worker.progress.connect(self._progress.setValue)
        self._worker.log.connect(self._log.append)
        self._worker.done.connect(self._on_finished)
        self._worker.start()

    def _on_finished(self, result):
        self._export_btn.setEnabled(True)
        msg = ("Exported " + str(result.get("exported", 0)) + " folder(s), skipped "
               + str(result.get("skipped", 0)) + ".")
        self._log.append(msg)
        if result.get("error"):
            QMessageBox.warning(self, "Export finished with error",
                                msg + "\n\nError: " + str(result["error"]))
        else:
            QMessageBox.information(self, "Export complete", msg)

    def closeEvent(self, event):
        if self._worker is not None and self._worker.isRunning():
            self._worker.wait()
        super().closeEvent(event)


def main():
    app = QApplication(sys.argv)
    dlg = ExportDialog()
    dlg.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
